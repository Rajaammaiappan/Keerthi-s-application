"""
OptiView -- FTE Location & Workforce Planning Tool (standalone build)

Self-contained: the entire app -- every route, every HTML template, and
the CSS/JS -- is baked into this one file, so it runs with nothing else
alongside it, just the pip-installed dependencies below.

GENERATED FILE -- do not hand-edit. Edit the source under app/ and run
build_standalone.py again instead; it regenerates this file from scratch.

Setup:
    pip install fastapi "uvicorn[standard]" pandas openpyxl python-multipart jinja2 xlrd numpy
Run:
    python OptiView_Standalone.py
Then open:
    http://127.0.0.1:8000
"""
from __future__ import annotations
import json
from jinja2 import Environment, DictLoader
TEMPLATES = {'analysis.html': '{% extends "base.html" %}\n{% block title %}FTE Analysis | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div>\n    <h1>FTE Analysis</h1>\n    <p class="subtitle">Compare periods and view location capacity/demand.</p>\n  </div>\n  <a href="#" id="export-link" class="btn btn-primary">Download Analyzed Report (.xlsx)</a>\n</section>\n\n<section id="filter-bar" class="filter-bar" data-periods=\'{{ periods|tojson }}\' data-categories=\'{{ categories|tojson }}\' data-dims=\'{{ dims|tojson }}\'></section>\n\n<section class="compare-bar">\n  <label>From Period\n    <select id="from-period"></select>\n  </label>\n  <label>To Period\n    <select id="to-period"></select>\n  </label>\n</section>\n\n<section class="panel">\n  <h2>Period Comparison</h2>\n  <div class="table-scroll"><table id="comparison-table" class="data-table"></table></div>\n</section>\n\n<section class="panel">\n  <h2>Location Capacity / Demand</h2>\n  <div class="table-scroll"><table id="capacity-table" class="data-table"></table></div>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initAnalysis());</script>\n{% endblock %}\n', 'base.html': '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">\n<meta name="viewport" content="width=device-width, initial-scale=1.0">\n<title>{% block title %}OptiView{% endblock %}</title>\n<link rel="stylesheet" href="/static/css/style.css?v=6">\n</head>\n<body>\n<header class="topbar">\n  <div class="topbar-inner">\n    <a class="brand" href="/">OptiView</a>\n    <nav class="navlinks">\n      <a href="/" class="{% if request.url.path == \'/\' %}active{% endif %}">Home</a>\n      <a href="/dashboard" class="{% if request.url.path == \'/dashboard\' %}active{% endif %}">Dashboard</a>\n      <a href="/mapping" class="{% if request.url.path == \'/mapping\' %}active{% endif %}">Location Mapping</a>\n      <a href="/analysis" class="{% if request.url.path == \'/analysis\' %}active{% endif %}">FTE Analysis</a>\n      <a href="/transfer" class="{% if request.url.path == \'/transfer\' %}active{% endif %}">Transfer Mapping</a>\n      <a href="/data-quality" class="{% if request.url.path == \'/data-quality\' %}active{% endif %}">Data Quality</a>\n      <a href="/management" class="{% if request.url.path == \'/management\' %}active{% endif %}">Management View</a>\n    </nav>\n  </div>\n</header>\n<main class="page">\n{% block content %}{% endblock %}\n</main>\n<footer class="footer">OptiView — FTE Location &amp; Workforce Planning Tool</footer>\n\n<div id="assistant-widget" class="assistant-widget" hidden>\n  <div id="assistant-panel" class="assistant-panel" hidden>\n    <div class="assistant-panel-header">\n      <span>🤖 OptiView Assistant</span>\n      <button id="assistant-close" type="button" aria-label="Close">&times;</button>\n    </div>\n    <div class="assistant-panel-body">\n      <div id="assistant-widget-feed" class="assistant-widget-feed">\n        <div class="assistant-hint">\n          <p>Try:</p>\n          <ul>\n            <li>"Total FTE", "top location", "top customer"</li>\n            <li>"Base issues", "negative FTE"</li>\n            <li>"Releasing capacity", "growing locations"</li>\n            <li>"Transfer opportunities", "category split"</li>\n            <li>Or just type a customer or location name.</li>\n          </ul>\n        </div>\n      </div>\n      <div id="assistant-builder" class="assistant-builder" hidden>\n        <label>Metric\n          <select id="assistant-metric-select"></select>\n        </label>\n        <div class="assistant-builder-periods">\n          <label>Period <select id="assistant-period"></select></label>\n          <label>From <select id="assistant-from-period"></select></label>\n          <label>To <select id="assistant-to-period"></select></label>\n        </div>\n        <button id="assistant-builder-ask" class="btn btn-primary btn-block" type="button">Ask</button>\n      </div>\n    </div>\n    <div id="assistant-quick-pills" class="assistant-quick-pills"></div>\n    <form id="assistant-input-form" class="assistant-input-row">\n      <input id="assistant-input" type="text" placeholder="Ask OptiView anything…" autocomplete="off">\n      <button type="submit" class="assistant-send" aria-label="Send">&#10148;</button>\n    </form>\n  </div>\n  <button id="assistant-fab" class="assistant-fab" type="button" aria-label="Open OptiView Assistant">💬</button>\n</div>\n\n<script src="/static/js/app.js?v=4"></script>\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initAssistantWidget());</script>\n{% block scripts %}{% endblock %}\n</body>\n</html>\n', 'dashboard.html': '{% extends "base.html" %}\n{% block title %}Dashboard | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div>\n    <h1>Dashboard</h1>\n    <p class="subtitle">Source file: <strong>{{ filename }}</strong></p>\n  </div>\n  <form action="/reset" method="post"><button class="btn btn-outline" type="submit">Reset / Upload New File</button></form>\n</section>\n\n{% if warnings %}\n<div class="warning-box">\n  <strong>Data quality notices:</strong>\n  <ul>{% for w in warnings %}<li>{{ w }}</li>{% endfor %}</ul>\n  <a href="/data-quality">View full Data Quality report →</a>\n</div>\n{% endif %}\n\n<section id="filter-bar" class="filter-bar" data-periods=\'{{ periods|tojson }}\' data-categories=\'{{ categories|tojson }}\' data-dims=\'{{ dims|tojson }}\'></section>\n\n<section class="summary-cards" id="summary-cards"></section>\n\n<section class="grid-2">\n  <div class="panel">\n    <div class="panel-header"><h2>FTE by Category</h2></div>\n    <div id="category-pie"></div>\n  </div>\n  <div class="panel">\n    <div class="panel-header"><h2>FTE by Location — Internal / SWC / External / Others</h2></div>\n    <div id="location-bar"></div>\n  </div>\n</section>\n\n<section class="panel">\n  <div class="panel-header">\n    <h2>Location Mapping Matrix (preview)</h2>\n    <a href="/mapping" class="btn btn-link">Open full matrix →</a>\n  </div>\n  <div class="table-scroll"><table id="matrix-table" class="matrix-table"></table></div>\n</section>\n\n<section class="grid-2">\n  <div class="panel">\n    <h2>Location Summary</h2>\n    <div class="table-scroll"><table id="location-summary-table" class="data-table"></table></div>\n  </div>\n  <div class="panel">\n    <h2>Customer Summary</h2>\n    <div class="table-scroll"><table id="customer-summary-table" class="data-table"></table></div>\n  </div>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initDashboard());</script>\n{% endblock %}\n', 'data_quality.html': '{% extends "base.html" %}\n{% block title %}Data Quality | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div><h1>Data Quality</h1><p class="subtitle">Validation findings from the uploaded workbook.</p></div>\n</section>\n\n<section class="panel">\n  <h2>Parsing Warnings</h2>\n  {% if report.parse_warnings %}\n  <ul>{% for w in report.parse_warnings %}<li>{{ w }}</li>{% endfor %}</ul>\n  {% else %}<p class="muted">No parsing warnings.</p>{% endif %}\n</section>\n\n<section class="grid-2">\n  <div class="panel">\n    <h2>⚠ Base Location Not Defined</h2>\n    {% if report.missing_base_customers %}\n    <ul>{% for c in report.missing_base_customers %}<li>{{ c }}</li>{% endfor %}</ul>\n    {% else %}<p class="muted">All accounts have a defined base location.</p>{% endif %}\n  </div>\n  <div class="panel">\n    <h2>⚠ Multiple Base Locations</h2>\n    {% if report.multiple_base_customers %}\n    <ul>{% for c in report.multiple_base_customers %}<li>{{ c }}</li>{% endfor %}</ul>\n    {% else %}<p class="muted">No accounts have multiple base locations.</p>{% endif %}\n  </div>\n</section>\n\n<section class="panel">\n  <h2>Negative FTE Records</h2>\n  {% if report.negative_fte_records %}\n  <div class="table-scroll">\n    <table class="data-table">\n      <thead><tr><th>Customer</th><th>Location</th><th>Period</th><th>Category</th><th>FTE</th></tr></thead>\n      <tbody>\n        {% for r in report.negative_fte_records %}\n        <tr><td>{{ r.Customer_Account }}</td><td>{{ r.Location }}</td><td>{{ r.Period }}</td><td>{{ r.Category }}</td><td class="negative">{{ r.FTE }}</td></tr>\n        {% endfor %}\n      </tbody>\n    </table>\n  </div>\n  {% else %}<p class="muted">No negative FTE values found.</p>{% endif %}\n</section>\n{% endblock %}\n', 'error.html': '{% extends "base.html" %}\n{% block title %}Error | OptiView{% endblock %}\n{% block content %}\n<section class="hero">\n  <h1>Something went wrong</h1>\n  <div class="error-box">\n    <p>{{ message }}</p>\n    {% if detail %}<p class="muted">{{ detail }}</p>{% endif %}\n  </div>\n  <a href="/" class="btn btn-primary">Back to Upload</a>\n</section>\n{% endblock %}\n', 'index.html': '{% extends "base.html" %}\n{% block title %}Upload | OptiView{% endblock %}\n{% block content %}\n<section class="hero">\n  <h1>Upload Workforce / FTE Excel</h1>\n  <p class="subtitle">\n    Upload your workforce/FTE tracking Excel file. The tool will automatically detect\n    customer accounts, locations, base locations, and FTE periods/categories, then build\n    a Location Mapping Matrix and workforce planning dashboard.\n  </p>\n\n  {% if has_dataset %}\n  <div class="notice">\n    A dataset is already loaded (<strong>{{ filename }}</strong>).\n    <a href="/dashboard">Go to Dashboard →</a>\n  </div>\n  {% endif %}\n\n  <form id="upload-form" action="/upload" method="post" enctype="multipart/form-data">\n    <label for="file-input" class="dropzone" id="dropzone">\n      <input type="file" id="file-input" name="file" accept=".xlsx,.xls" required>\n      <svg class="dropzone-icon" viewBox="0 0 24 24" fill="none" aria-hidden="true">\n        <path d="M12 4v11m0-11 4 4m-4-4-4 4" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>\n        <path d="M4 16v2.5A2.5 2.5 0 0 0 6.5 21h11a2.5 2.5 0 0 0 2.5-2.5V16" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>\n      </svg>\n      <div class="dropzone-text">\n        <span class="dropzone-title">Drag &amp; drop your Excel file here</span>\n        <span class="dropzone-sub">or click to browse (.xlsx, .xls)</span>\n        <span class="dropzone-filename" id="dropzone-filename"></span>\n      </div>\n    </label>\n    <button type="submit" class="btn btn-primary btn-block">Upload &amp; Analyze</button>\n  </form>\n\n  <div class="info-box">\n    <h3>Expected columns</h3>\n    <p>S.No, VM Product, Customer Account, Component, Country, Location, Base location,\n      and per-period FTE columns broken down as\n      <code>Internal</code>, <code>SWC</code>, <code>External</code>, <code>Others</code>,\n      and the period total. Any number of periods, categories, and locations is\n      detected automatically — column order doesn\'t matter.</p>\n  </div>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>\nconst dz = document.getElementById(\'dropzone\');\nconst input = document.getElementById(\'file-input\');\nconst nameLabel = document.getElementById(\'dropzone-filename\');\ninput.addEventListener(\'change\', () => {\n  nameLabel.textContent = input.files.length ? input.files[0].name : \'\';\n});\n[\'dragover\',\'dragenter\'].forEach(evt => dz.addEventListener(evt, e => { e.preventDefault(); dz.classList.add(\'dragover\'); }));\n[\'dragleave\',\'drop\'].forEach(evt => dz.addEventListener(evt, e => { e.preventDefault(); dz.classList.remove(\'dragover\'); }));\ndz.addEventListener(\'drop\', e => {\n  if (e.dataTransfer.files.length) {\n    input.files = e.dataTransfer.files;\n    nameLabel.textContent = input.files[0].name;\n  }\n});\n</script>\n{% endblock %}\n', 'management.html': '{% extends "base.html" %}\n{% block title %}Management View | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div><h1>Management View</h1><p class="subtitle">High-level summary — no technical detail.</p></div>\n  <a href="#" id="export-link" class="btn btn-primary">Download Report</a>\n</section>\n\n<section id="filter-bar" class="filter-bar" data-periods=\'{{ periods|tojson }}\' data-categories=\'{{ categories|tojson }}\' data-dims=\'{{ dims|tojson }}\' data-simple="1"></section>\n\n<section class="summary-cards" id="summary-cards"></section>\n\n<section class="grid-2">\n  <div class="panel">\n    <h2>Location Distribution</h2>\n    <div class="table-scroll"><table id="location-summary-table" class="data-table"></table></div>\n  </div>\n  <div class="panel">\n    <h2>Customer Distribution</h2>\n    <div class="table-scroll"><table id="customer-summary-table" class="data-table"></table></div>\n  </div>\n</section>\n\n<section class="compare-bar">\n  <label>From Period<select id="from-period"></select></label>\n  <label>To Period<select id="to-period"></select></label>\n</section>\n\n<section class="panel">\n  <h2>FTE Growth / Reduction</h2>\n  <div class="table-scroll"><table id="comparison-table" class="data-table"></table></div>\n</section>\n\n<section class="panel">\n  <h2>Top Potential Transfer Opportunities</h2>\n  <div class="table-scroll"><table id="transfer-table" class="data-table"></table></div>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initManagement());</script>\n{% endblock %}\n', 'mapping.html': '{% extends "base.html" %}\n{% block title %}Location Mapping | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div>\n    <h1>Location Mapping Matrix</h1>\n    <p class="subtitle">Customer Account × Location — base locations, supporting locations and FTE.</p>\n  </div>\n  <a href="#" id="export-link" class="btn btn-primary">Download Analyzed Report (.xlsx)</a>\n</section>\n\n<section id="filter-bar" class="filter-bar" data-periods=\'{{ periods|tojson }}\' data-categories=\'{{ categories|tojson }}\' data-dims=\'{{ dims|tojson }}\'></section>\n\n<div class="view-toggle" id="view-toggle">\n  <button data-view="status" class="active">★ / ✓ Status</button>\n  <button data-view="fte">FTE Numbers</button>\n  <button data-view="both">Status + FTE</button>\n  <button data-view="breakdown">Internal / SWC / External Breakdown</button>\n</div>\n\n<section class="panel">\n  <div class="table-scroll"><table id="matrix-table" class="matrix-table"></table></div>\n  <div id="matrix-category-legend"></div>\n</section>\n\n<section class="legend">\n  <span><strong>★ BASE</strong> — designated base location</span>\n  <span><strong>✓</strong> — supporting / additional location</span>\n  <span><strong>-</strong> — no FTE at this location</span>\n  <span class="legend-warn">⚠ — base-location data-quality issue (see Data Quality page)</span>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initMapping());</script>\n{% endblock %}\n', 'transfer.html': '{% extends "base.html" %}\n{% block title %}Transfer Mapping | OptiView{% endblock %}\n{% block content %}\n<section class="page-header">\n  <div>\n    <h1>Transfer Mapping</h1>\n    <p class="subtitle">Potential transfer opportunities between locations, and manual mapping entries.</p>\n  </div>\n  <a href="#" id="export-link" class="btn btn-primary">Download Analyzed Report (.xlsx)</a>\n</section>\n\n<section id="filter-bar" class="filter-bar" data-periods=\'{{ periods|tojson }}\' data-categories=\'{{ categories|tojson }}\' data-dims=\'{{ dims|tojson }}\'></section>\n\n<section class="compare-bar">\n  <label>From Period\n    <select id="from-period"></select>\n  </label>\n  <label>To Period\n    <select id="to-period"></select>\n  </label>\n</section>\n\n<section class="panel">\n  <h2>Potential Transfer Opportunities</h2>\n  <p class="muted">Locations releasing capacity matched against locations/accounts with growing demand.</p>\n  <div class="table-scroll"><table id="transfer-table" class="data-table"></table></div>\n</section>\n\n<section class="panel">\n  <h2>Create Manual Mapping</h2>\n  <form id="manual-mapping-form" class="inline-form">\n    <input name="from_customer" placeholder="From Customer" required>\n    <input name="from_location" placeholder="From Location" required>\n    <input name="to_customer" placeholder="To Customer" required>\n    <input name="to_location" placeholder="To Location" required>\n    <input name="fte_amount" type="number" step="0.1" placeholder="FTE Amount" required>\n    <input name="notes" placeholder="Notes (optional)">\n    <button type="submit" class="btn btn-primary">Add Mapping</button>\n  </form>\n  <div class="table-scroll"><table id="manual-mapping-table" class="data-table"></table></div>\n</section>\n{% endblock %}\n{% block scripts %}\n<script>document.addEventListener(\'DOMContentLoaded\', () => FTE.initTransfer());</script>\n{% endblock %}\n'}
STYLE_CSS = ':root {\n  --navy: #1f3864;\n  --navy-dark: #14284a;\n  --accent: #2f6fed;\n  --accent-dark: #2559c4;\n  --bg: #f4f6fa;\n  --panel-bg: #ffffff;\n  --border: #e1e5eb;\n  --text: #1e2430;\n  --muted: #6b7380;\n  --success: #1a7f4b;\n  --warn: #b8860b;\n  --danger: #c0392b;\n  --base-bg: #fff2cc;\n  --support-bg: #e7f0ff;\n  --shadow-sm: 0 1px 2px rgba(20, 30, 60, 0.06);\n  --shadow-md: 0 4px 14px rgba(20, 30, 60, 0.08);\n  /* FTE category colors -- shared by the pie chart, bar chart, matrix\n     breakdown cells and every legend, so a category always reads the\n     same color everywhere in the app. */\n  --cat-internal: #2f6fed;\n  --cat-swc: #8b5cf6;\n  --cat-external: #f59e0b;\n  --cat-others: #64748b;\n}\n\n* { box-sizing: border-box; }\nbody {\n  margin: 0;\n  font-family: "Segoe UI", Roboto, Arial, sans-serif;\n  background: var(--bg);\n  color: var(--text);\n  font-size: 14px;\n  -webkit-font-smoothing: antialiased;\n}\n\n.topbar { background: var(--navy); color: #fff; box-shadow: 0 2px 8px rgba(20, 30, 60, 0.15); position: sticky; top: 0; z-index: 10; }\n.topbar-inner {\n  max-width: 1280px; margin: 0 auto; padding: 0 20px;\n  display: flex; align-items: center; justify-content: space-between; height: 56px;\n  flex-wrap: wrap;\n}\n.brand { color: #fff; font-weight: 600; text-decoration: none; font-size: 16px; letter-spacing: .01em; }\n.navlinks { display: flex; gap: 2px; flex-wrap: wrap; }\n.navlinks a {\n  color: #cfd8e8; text-decoration: none; padding: 8px 12px; border-radius: 5px; font-size: 13px;\n  transition: background-color .15s, color .15s;\n}\n.navlinks a:hover { background: rgba(255,255,255,.08); color: #fff; }\n.navlinks a.active { background: var(--navy-dark); color: #fff; font-weight: 600; }\n\n.page { max-width: 1280px; margin: 0 auto; padding: 24px 20px 60px; }\n.footer { text-align: center; color: var(--muted); font-size: 12px; padding: 20px; }\n\nh1 { font-size: 22px; margin: 0 0 4px; letter-spacing: -.01em; }\nh2 { font-size: 16px; margin: 0 0 12px; color: var(--navy); }\n.subtitle { color: var(--muted); margin: 0; }\n.muted { color: var(--muted); }\n\n.page-header {\n  display: flex; justify-content: space-between; align-items: flex-start;\n  margin-bottom: 16px; gap: 12px; flex-wrap: wrap;\n}\n\n.hero { max-width: 480px; margin: 48px auto; text-align: center; }\n.hero h1 { font-size: 26px; }\n\n.btn {\n  display: inline-block; padding: 9px 16px; border-radius: 6px; border: 1px solid transparent;\n  font-size: 13px; font-weight: 600; cursor: pointer; text-decoration: none; text-align: center;\n  transition: background-color .15s, border-color .15s, box-shadow .15s, transform .05s;\n}\n.btn:active { transform: translateY(1px); }\n.btn-primary { background: var(--accent); color: #fff; box-shadow: var(--shadow-sm); }\n.btn-primary:hover { background: var(--accent-dark); }\n.btn-secondary { background: var(--navy); color: #fff; }\n.btn-secondary:hover { background: var(--navy-dark); }\n.btn-outline { background: #fff; color: var(--navy); border-color: var(--border); }\n.btn-outline:hover { border-color: var(--navy); }\n.btn-link { background: none; color: var(--accent); border: none; padding: 9px 4px; }\n.btn-block { display: block; width: 100%; padding: 11px 16px; font-size: 14px; }\n\n.dropzone {\n  display: flex; flex-direction: column; align-items: center; text-align: center;\n  border: 2px dashed var(--border); border-radius: 10px; padding: 44px 20px;\n  background: var(--panel-bg); cursor: pointer; margin: 20px 0;\n  transition: border-color .15s, background-color .15s;\n}\n.dropzone:hover { border-color: #b9c6dc; }\n.dropzone.dragover { border-color: var(--accent); background: #eef4ff; }\n.dropzone input[type=file] { display: none; }\n.dropzone-icon { width: 30px; height: 30px; color: var(--accent); margin-bottom: 12px; }\n.dropzone-title { display: block; font-weight: 600; font-size: 15px; }\n.dropzone-sub { display: block; color: var(--muted); margin-top: 4px; font-size: 12px; }\n.dropzone-filename { display: block; margin-top: 10px; color: var(--accent); font-weight: 600; }\n\n.info-box { text-align: left; background: var(--panel-bg); border: 1px solid var(--border); border-radius: 8px; padding: 16px; margin-top: 24px; box-shadow: var(--shadow-sm); }\n.info-box h3 { margin: 0 0 6px; font-size: 13px; color: var(--navy); }\n.info-box p { margin: 0; color: var(--muted); line-height: 1.6; }\n.info-box code { background: #eef1f5; padding: 1px 5px; border-radius: 3px; color: var(--text); }\n\n.notice { background: #eaf7ee; border: 1px solid #bfe6c9; padding: 10px 14px; border-radius: 6px; margin-bottom: 16px; }\n.warning-box { background: #fff8e6; border: 1px solid #f0d998; padding: 12px 16px; border-radius: 6px; margin-bottom: 16px; }\n.warning-box ul { margin: 6px 0 0 18px; }\n.error-box { background: #fdecea; border: 1px solid #f5c6c1; padding: 16px; border-radius: 6px; margin: 16px 0; }\n\n.filter-bar {\n  display: flex; gap: 14px; flex-wrap: wrap; align-items: flex-end;\n  background: var(--panel-bg); border: 1px solid var(--border); border-radius: 8px;\n  padding: 14px 16px; margin-bottom: 18px; box-shadow: var(--shadow-sm);\n}\n.filter-field { display: flex; flex-direction: column; gap: 4px; font-size: 12px; color: var(--muted); min-width: 140px; }\n.filter-field select, .filter-field input {\n  padding: 6px 8px; border: 1px solid var(--border); border-radius: 5px; font-size: 13px;\n  background: #fff; transition: border-color .15s;\n}\n.filter-field select:focus, .filter-field input:focus { outline: none; border-color: var(--accent); }\n\n.summary-cards { display: grid; grid-template-columns: repeat(6, 1fr); gap: 12px; margin-bottom: 20px; }\n.card {\n  background: var(--panel-bg); border: 1px solid var(--border); border-radius: 8px; padding: 14px 16px;\n  box-shadow: var(--shadow-sm); transition: box-shadow .15s, transform .15s;\n}\n.card:hover { box-shadow: var(--shadow-md); transform: translateY(-1px); }\n.card .card-label { font-size: 11px; text-transform: uppercase; letter-spacing: .04em; color: var(--muted); }\n.card .card-value { font-size: 24px; font-weight: 700; color: var(--navy); margin-top: 4px; }\n\n.panel { background: var(--panel-bg); border: 1px solid var(--border); border-radius: 8px; padding: 18px; margin-bottom: 20px; box-shadow: var(--shadow-sm); }\n.panel-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }\n.grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }\n\n.table-scroll { overflow-x: auto; border-radius: 6px; }\ntable { border-collapse: collapse; width: 100%; font-size: 13px; }\nth, td { padding: 8px 10px; border-bottom: 1px solid var(--border); text-align: left; white-space: nowrap; }\nth { background: #f0f2f5; color: var(--navy); position: sticky; top: 0; font-weight: 600; }\ntbody tr:hover { background: #f7f9fc; }\n\n.matrix-table td.base-cell { background: var(--base-bg); font-weight: 700; text-align: center; }\n.matrix-table td.support-cell { background: var(--support-bg); text-align: center; }\n.matrix-table td.empty-cell { text-align: center; color: var(--muted); }\n.matrix-table td.negative-cell { background: #fbe3e0; color: var(--danger); font-weight: 700; text-align: center; }\n.matrix-table td.fte-sub { display: block; font-size: 11px; font-weight: 400; color: var(--muted); }\n.matrix-table th.customer-col-warn { color: var(--danger); }\n\n.matrix-table td.breakdown-cell { text-align: center; padding: 6px 8px; }\n.breakdown-line { white-space: nowrap; font-size: 12px; font-weight: 700; }\n.breakdown-line .op { color: var(--muted); font-weight: 400; margin: 0 2px; }\n.breakdown-total { font-size: 11px; color: var(--muted); margin-top: 2px; }\n.breakdown-total strong { color: var(--text); }\n\n/* -------------------------------------------------------------------\n   Category legend -- reused by the matrix breakdown view and the\n   Dashboard charts so Internal/SWC/External/Others always map to the\n   same swatch wherever they appear.\n-------------------------------------------------------------------- */\n.category-legend { display: flex; gap: 16px; flex-wrap: wrap; font-size: 12px; color: var(--text); margin-top: 12px; }\n.category-legend .legend-item { display: flex; align-items: center; gap: 6px; }\n.category-legend .legend-swatch { width: 10px; height: 10px; border-radius: 3px; flex: none; }\n\n/* -------------------------------------------------------------------\n   Charts -- hand-rolled SVG/CSS, no charting library.\n-------------------------------------------------------------------- */\n.chart-empty { color: var(--muted); font-size: 13px; padding: 20px 0; text-align: center; }\n\n.pie-chart-wrap { display: flex; align-items: center; gap: 24px; flex-wrap: wrap; }\n.pie-chart { width: 160px; height: 160px; flex: none; }\n.pie-chart-center {\n  position: relative; margin-left: -160px; width: 160px; height: 160px;\n  display: flex; flex-direction: column; align-items: center; justify-content: center;\n  pointer-events: none;\n}\n.pie-chart-center .pie-total { font-size: 20px; font-weight: 700; color: var(--navy); }\n.pie-chart-center .pie-total-label { font-size: 10px; color: var(--muted); text-transform: uppercase; letter-spacing: .04em; }\n.pie-chart-legend { display: flex; flex-direction: column; gap: 8px; font-size: 13px; }\n.pie-chart-legend .legend-item { display: flex; align-items: center; gap: 8px; }\n.pie-chart-legend .legend-swatch { width: 11px; height: 11px; border-radius: 3px; flex: none; }\n.pie-chart-legend .legend-value { margin-left: auto; font-weight: 600; color: var(--text); padding-left: 12px; }\n.pie-chart-legend .legend-pct { color: var(--muted); font-size: 11px; width: 40px; text-align: right; }\n\n.bar-chart { display: flex; flex-direction: column; gap: 10px; }\n.bar-row { display: grid; grid-template-columns: 70px 1fr 44px; align-items: center; gap: 10px; }\n.bar-label { font-size: 12px; color: var(--text); font-weight: 600; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }\n.bar-track { display: flex; height: 16px; background: #eef1f5; border-radius: 4px; overflow: hidden; }\n.bar-seg { height: 100%; transition: width .2s; }\n.bar-total { font-size: 12px; color: var(--muted); text-align: right; }\n\n/* -------------------------------------------------------------------\n   Chat bubbles -- shared by the assistant widget\'s feed.\n-------------------------------------------------------------------- */\n.chat-entry { display: flex; }\n.chat-entry.chat-question { justify-content: flex-end; }\n.chat-entry.chat-answer, .chat-entry.chat-error { flex-direction: column; align-items: flex-start; }\n.chat-bubble {\n  display: inline-block; padding: 10px 14px; border-radius: 12px; font-size: 13px; line-height: 1.5;\n  max-width: 85%;\n}\n.chat-question .chat-bubble { background: var(--navy); color: #fff; border-bottom-right-radius: 3px; }\n.chat-answer .chat-bubble { background: #eef1f5; color: var(--text); border-bottom-left-radius: 3px; }\n.chat-error .chat-bubble { background: #fdecea; color: var(--danger); border-bottom-left-radius: 3px; }\n.chat-table { margin-top: 8px; font-size: 12px; width: auto; min-width: 260px; }\n.chat-table th, .chat-table td { padding: 6px 10px; white-space: nowrap; }\n\n/* -------------------------------------------------------------------\n   Floating Assistant widget -- available on every page as a popup,\n   not a separate tab.\n-------------------------------------------------------------------- */\n.assistant-widget { position: fixed; right: 24px; bottom: 24px; z-index: 200; }\n\n.assistant-fab {\n  width: 56px; height: 56px; border-radius: 50%; border: none; background: var(--navy);\n  color: #fff; font-size: 24px; cursor: pointer; box-shadow: var(--shadow-md);\n  display: flex; align-items: center; justify-content: center; transition: transform .15s, background-color .15s;\n  margin-left: auto;\n}\n.assistant-fab:hover { transform: scale(1.06); background: var(--navy-dark); }\n\n.assistant-panel {\n  position: absolute; right: 0; bottom: 68px; width: 340px; max-height: 70vh;\n  background: var(--panel-bg); border-radius: 14px; box-shadow: var(--shadow-md);\n  display: flex; flex-direction: column; overflow: hidden; border: 1px solid var(--border);\n}\n/* `display: flex` above has the same specificity as the browser\'s default\n   `[hidden] { display: none }` rule and wins by source order, so the panel\n   never actually hides -- this rule (higher specificity: class + attribute)\n   forces it closed when the `hidden` attribute is set. */\n.assistant-panel[hidden] { display: none; }\n.assistant-panel-header {\n  background: var(--navy); color: #fff; padding: 12px 16px; display: flex;\n  align-items: center; justify-content: space-between; font-weight: 600; font-size: 14px; flex: none;\n}\n.assistant-panel-header button {\n  background: none; border: none; color: #fff; font-size: 20px; cursor: pointer; line-height: 1;\n  min-width: 28px; min-height: 28px; display: flex; align-items: center; justify-content: center;\n  border-radius: 6px;\n}\n.assistant-panel-header button:hover { background: rgba(255, 255, 255, .15); }\n.assistant-panel-body { flex: 1; overflow-y: auto; padding: 12px 14px; display: flex; flex-direction: column; gap: 10px; min-height: 120px; }\n.assistant-hint { font-size: 12px; color: var(--muted); }\n.assistant-hint p { margin: 0 0 4px; font-weight: 600; color: var(--text); }\n.assistant-hint ul { margin: 0 0 0 16px; padding: 0; }\n.assistant-hint li { margin-bottom: 4px; }\n\n.assistant-builder {\n  border: 1px solid var(--border); border-radius: 8px; padding: 10px; display: flex;\n  flex-direction: column; gap: 8px; background: #f7f9fc; font-size: 12px; flex: none;\n}\n.assistant-builder label { display: flex; flex-direction: column; gap: 3px; color: var(--muted); }\n.assistant-builder select { padding: 5px 6px; border: 1px solid var(--border); border-radius: 4px; font-size: 12px; }\n.assistant-builder-periods { display: flex; gap: 6px; }\n.assistant-builder-periods label { flex: 1; }\n\n.assistant-quick-pills { display: flex; flex-wrap: wrap; gap: 6px; padding: 10px 14px; border-top: 1px solid var(--border); flex: none; }\n.assistant-pill {\n  border: 1px solid var(--border); background: #fff; border-radius: 20px; padding: 6px 12px;\n  font-size: 12px; font-weight: 600; color: var(--navy); cursor: pointer; white-space: nowrap;\n  transition: border-color .15s, color .15s;\n}\n.assistant-pill:hover { border-color: var(--accent); color: var(--accent); }\n.assistant-pill-accent { background: #eef4ff; border-color: var(--accent); color: var(--accent-dark); }\n\n.assistant-input-row { display: flex; gap: 8px; padding: 10px 14px; border-top: 1px solid var(--border); flex: none; }\n.assistant-input-row input {\n  flex: 1; padding: 8px 12px; border: 1px solid var(--border); border-radius: 20px; font-size: 13px; min-width: 0;\n}\n.assistant-input-row input:focus { outline: none; border-color: var(--accent); }\n.assistant-send {\n  width: 36px; height: 36px; border-radius: 50%; border: none; background: var(--accent); color: #fff;\n  cursor: pointer; flex: none; font-size: 14px;\n}\n.assistant-send:hover { background: var(--accent-dark); }\n\n@media (max-width: 480px) {\n  .assistant-widget { right: 12px; bottom: 12px; }\n  .assistant-panel { width: calc(100vw - 24px); }\n}\n\n.view-toggle { display: flex; gap: 8px; margin-bottom: 14px; }\n.view-toggle button {\n  padding: 7px 14px; border: 1px solid var(--border); background: #fff; border-radius: 20px;\n  cursor: pointer; font-size: 12px; font-weight: 600; color: var(--muted);\n}\n.view-toggle button.active { background: var(--navy); color: #fff; border-color: var(--navy); }\n\n.legend { display: flex; gap: 20px; flex-wrap: wrap; font-size: 12px; color: var(--muted); }\n.legend-warn { color: var(--danger); }\n\n.compare-bar { display: flex; gap: 16px; margin-bottom: 16px; }\n.compare-bar label { display: flex; flex-direction: column; font-size: 12px; color: var(--muted); gap: 4px; }\n.compare-bar select { padding: 6px 8px; border: 1px solid var(--border); border-radius: 4px; }\n\n.status-growth { color: var(--success); font-weight: 600; }\n.status-reduction { color: var(--danger); font-weight: 600; }\n.status-new { color: var(--accent); font-weight: 600; }\n.status-discontinued { color: var(--muted); font-weight: 600; }\n.status-stable { color: var(--muted); }\n.negative { color: var(--danger); font-weight: 700; }\n\n.inline-form { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }\n.inline-form input { padding: 7px 9px; border: 1px solid var(--border); border-radius: 4px; font-size: 13px; }\n.inline-form input[name=notes] { flex: 1; min-width: 160px; }\n\n@media (max-width: 900px) {\n  .summary-cards { grid-template-columns: repeat(2, 1fr); }\n  .grid-2 { grid-template-columns: 1fr; }\n}\n'
APP_JS = '/* FTE Location & Workforce Planner — client-side logic.\n   Vanilla JS only, no frameworks. Talks to the JSON API under /api/*. */\n\nconst FTE = (() => {\n  const DIM_LABELS = {\n    Customer_Account: "Customer Account",\n    VM_Product: "VM Product",\n    Component: "Component",\n    Country: "Country",\n    Location: "Location",\n  };\n\n  // Fixed color per FTE category, used consistently by the pie chart, bar\n  // chart, matrix breakdown cells, and every legend.\n  const CATEGORY_COLORS = {\n    Internal: "#2f6fed",\n    SWC: "#8b5cf6",\n    External: "#f59e0b",\n    Others: "#64748b",\n  };\n  function categoryColor(cat) { return CATEGORY_COLORS[cat] || "#94a3b8"; }\n\n  const state = {\n    periods: [],\n    categories: [],\n    dims: {},\n    filters: {},\n    period: null,\n    category: "Total",\n    view: "status",\n    fromPeriod: null,\n    toPeriod: null,\n  };\n\n  function qs(sel, root = document) { return root.querySelector(sel); }\n  function qsa(sel, root = document) { return Array.from(root.querySelectorAll(sel)); }\n\n  async function api(path, params = {}) {\n    const url = new URL(path, window.location.origin);\n    for (const [key, val] of Object.entries(params)) {\n      if (val === undefined || val === null || val === "") continue;\n      if (Array.isArray(val)) {\n        val.forEach(v => url.searchParams.append(key, v));\n      } else {\n        url.searchParams.set(key, val);\n      }\n    }\n    const res = await fetch(url, { credentials: "same-origin" });\n    if (!res.ok) {\n      let message = `API error ${res.status}`;\n      try {\n        const body = await res.json();\n        if (body && body.error) message = body.error;\n      } catch (_) { /* body wasn\'t JSON -- keep the generic message */ }\n      throw new Error(message);\n    }\n    return res.json();\n  }\n\n  function currentFilterParams() {\n    return { ...state.filters };\n  }\n\n  // -------------------------------------------------------------------\n  // Filter bar\n  // -------------------------------------------------------------------\n  function buildFilterBar(onChange) {\n    const bar = qs("#filter-bar");\n    if (!bar) return;\n    state.periods = JSON.parse(bar.dataset.periods || "[]");\n    state.categories = JSON.parse(bar.dataset.categories || "[]");\n    state.dims = JSON.parse(bar.dataset.dims || "{}");\n    state.period = state.periods[state.periods.length - 1] || null;\n    state.category = state.categories.includes("Total") ? "Total" : (state.categories[0] || "Total");\n    state.fromPeriod = state.periods[0] || null;\n    state.toPeriod = state.periods[state.periods.length - 1] || null;\n\n    bar.innerHTML = "";\n\n    // Period selector\n    if (!bar.dataset.simple) {\n      bar.appendChild(makeSelectField("Period", state.periods, state.period, val => {\n        state.period = val; onChange();\n      }));\n    }\n    // Category selector\n    bar.appendChild(makeSelectField("FTE Type", state.categories, state.category, val => {\n      state.category = val; onChange();\n    }));\n    // Dimension filters\n    for (const [dim, values] of Object.entries(state.dims)) {\n      bar.appendChild(makeMultiSelectField(DIM_LABELS[dim] || dim, dim, values, onChange));\n    }\n\n    // Populate from/to period selects if present on this page\n    const fromSel = qs("#from-period");\n    const toSel = qs("#to-period");\n    if (fromSel && toSel) {\n      fromSel.innerHTML = state.periods.map(p => `<option value="${p}">${p}</option>`).join("");\n      toSel.innerHTML = state.periods.map(p => `<option value="${p}">${p}</option>`).join("");\n      fromSel.value = state.fromPeriod;\n      toSel.value = state.toPeriod;\n      fromSel.addEventListener("change", () => { state.fromPeriod = fromSel.value; onChange(); });\n      toSel.addEventListener("change", () => { state.toPeriod = toSel.value; onChange(); });\n    }\n  }\n\n  function makeSelectField(label, options, current, onChange) {\n    const wrap = document.createElement("div");\n    wrap.className = "filter-field";\n    const lbl = document.createElement("label");\n    lbl.textContent = label;\n    const sel = document.createElement("select");\n    options.forEach(opt => {\n      const o = document.createElement("option");\n      o.value = opt; o.textContent = opt;\n      if (opt === current) o.selected = true;\n      sel.appendChild(o);\n    });\n    sel.addEventListener("change", () => onChange(sel.value));\n    wrap.appendChild(lbl); wrap.appendChild(sel);\n    return wrap;\n  }\n\n  function makeMultiSelectField(label, dim, values, onChange) {\n    const wrap = document.createElement("div");\n    wrap.className = "filter-field";\n    const lbl = document.createElement("label");\n    lbl.textContent = label + " (Ctrl/Cmd+click for multiple)";\n    const sel = document.createElement("select");\n    sel.multiple = true;\n    sel.size = Math.min(4, Math.max(2, values.length > 6 ? 4 : values.length));\n    values.forEach(v => {\n      const o = document.createElement("option");\n      o.value = v; o.textContent = v;\n      sel.appendChild(o);\n    });\n    sel.addEventListener("change", () => {\n      const selected = Array.from(sel.selectedOptions).map(o => o.value);\n      if (selected.length) state.filters[dim] = selected;\n      else delete state.filters[dim];\n      onChange();\n    });\n    wrap.appendChild(lbl); wrap.appendChild(sel);\n    return wrap;\n  }\n\n  // -------------------------------------------------------------------\n  // Matrix rendering\n  // -------------------------------------------------------------------\n  function renderMatrix(matrix) {\n    const table = qs("#matrix-table");\n    if (!table) return;\n    if (!matrix.customers.length || !matrix.locations.length) {\n      table.innerHTML = "<tr><td>No data for the current filters.</td></tr>";\n      return;\n    }\n    let html = "<thead><tr><th>Customer Account</th>";\n    matrix.locations.forEach(loc => { html += `<th>${loc}</th>`; });\n    html += "</tr></thead><tbody>";\n\n    matrix.customers.forEach(cust => {\n      const warning = matrix.customer_warnings[cust];\n      const custLabel = warning ? `⚠ ${cust}` : cust;\n      html += `<tr><th class="${warning ? \'customer-col-warn\' : \'\'}" title="${warning || \'\'}">${custLabel}</th>`;\n      matrix.locations.forEach(loc => {\n        const cell = matrix.cells[cust][loc];\n        html += renderCell(cell);\n      });\n      html += "</tr>";\n    });\n    html += "</tbody>";\n    table.innerHTML = html;\n  }\n\n  function renderCell(cell) {\n    const view = state.view || "status";\n    const fteVal = cell.fte;\n    const isNegative = fteVal !== null && fteVal < 0;\n    let cls = "empty-cell";\n    let content = "-";\n\n    if (isNegative) {\n      cls = "negative-cell";\n      content = view === "status" ? "⚠" : `${fteVal}`;\n      if (view === "both") content = `⚠<span class="fte-sub">${fteVal} FTE</span>`;\n    } else if (cell.is_base) {\n      cls = "base-cell";\n      if (view === "status") content = "★ BASE";\n      else if (view === "fte") content = cell.has_fte ? `${fteVal}` : "-";\n      else content = `★ BASE${cell.has_fte ? `<span class="fte-sub">${fteVal} FTE</span>` : ""}`;\n    } else if (cell.has_fte) {\n      cls = "support-cell";\n      if (view === "status") content = "✓";\n      else if (view === "fte") content = `${fteVal}`;\n      else content = `✓<span class="fte-sub">${fteVal} FTE</span>`;\n    }\n    return `<td class="${cls}">${content}</td>`;\n  }\n\n  // -------------------------------------------------------------------\n  // Matrix breakdown view (Internal + SWC + External + Others = Total,\n  // color-coded per category)\n  // -------------------------------------------------------------------\n  function renderMatrixBreakdown(matrix) {\n    const table = qs("#matrix-table");\n    const legendEl = qs("#matrix-category-legend");\n    if (!table) return;\n    if (!matrix.customers.length || !matrix.locations.length) {\n      table.innerHTML = "<tr><td>No data for the current filters.</td></tr>";\n      if (legendEl) legendEl.innerHTML = "";\n      return;\n    }\n    if (!matrix.categories.length) {\n      table.innerHTML = "<tr><td>This workbook has no Internal/SWC/External/Others breakdown — only period totals.</td></tr>";\n      if (legendEl) legendEl.innerHTML = "";\n      return;\n    }\n\n    let html = "<thead><tr><th>Customer Account</th>";\n    matrix.locations.forEach(loc => { html += `<th>${loc}</th>`; });\n    html += "</tr></thead><tbody>";\n\n    matrix.customers.forEach(cust => {\n      const warning = matrix.customer_warnings[cust];\n      const custLabel = warning ? `⚠ ${cust}` : cust;\n      html += `<tr><th class="${warning ? \'customer-col-warn\' : \'\'}" title="${warning || \'\'}">${custLabel}</th>`;\n      matrix.locations.forEach(loc => {\n        const cell = matrix.cells[cust][loc];\n        html += renderBreakdownCell(cell, matrix.categories);\n      });\n      html += "</tr>";\n    });\n    html += "</tbody>";\n    table.innerHTML = html;\n    if (legendEl) legendEl.innerHTML = renderCategoryLegend(matrix.categories);\n  }\n\n  function renderBreakdownCell(cell, categories) {\n    if (!cell.has_fte) return `<td class="empty-cell">-</td>`;\n\n    const isNegative = cell.total < 0;\n    const parts = categories\n      .map(cat => `<span style="color:${categoryColor(cat)}">${cell.values[cat]}</span>`)\n      .join(\'<span class="op">+</span>\');\n    const cls = isNegative ? "negative-cell" : cell.is_base ? "base-cell" : "support-cell";\n    return `<td class="${cls} breakdown-cell">\n      <div class="breakdown-line">${parts}</div>\n      <div class="breakdown-total">= <strong>${cell.total}</strong></div>\n    </td>`;\n  }\n\n  function renderCategoryLegend(categories) {\n    if (!categories.length) return "";\n    return `<div class="category-legend">${categories.map(cat =>\n      `<span class="legend-item"><span class="legend-swatch" style="background:${categoryColor(cat)}"></span>${cat}</span>`\n    ).join("")}</div>`;\n  }\n\n  // -------------------------------------------------------------------\n  // Charts (hand-rolled SVG/CSS -- no charting library dependency)\n  // -------------------------------------------------------------------\n  function renderCategoryPie(elId, breakdown) {\n    const el = qs(elId);\n    if (!el) return;\n    const data = (breakdown.by_category || []).filter(d => d.fte > 0);\n    const total = data.reduce((s, d) => s + d.fte, 0);\n    if (!data.length || total <= 0) {\n      el.innerHTML = \'<div class="chart-empty">No category breakdown for the current filters.</div>\';\n      return;\n    }\n\n    const cx = 80, cy = 80, r = 78, hole = 46;\n    let angle = -90;\n    let slices = "";\n    data.forEach(d => {\n      const frac = d.fte / total;\n      const sweep = Math.max(frac * 360, 0.001);\n      const a1 = angle * Math.PI / 180;\n      const a2 = (angle + sweep) * Math.PI / 180;\n      const x1 = cx + r * Math.cos(a1), y1 = cy + r * Math.sin(a1);\n      const x2 = cx + r * Math.cos(a2), y2 = cy + r * Math.sin(a2);\n      const largeArc = sweep > 180 ? 1 : 0;\n      slices += `<path d="M${cx},${cy} L${x1.toFixed(2)},${y1.toFixed(2)} A${r},${r} 0 ${largeArc} 1 ${x2.toFixed(2)},${y2.toFixed(2)} Z" fill="${categoryColor(d.category)}"><title>${d.category}: ${d.fte} (${(frac * 100).toFixed(1)}%)</title></path>`;\n      angle += sweep;\n    });\n    const svg = `<svg viewBox="0 0 160 160" class="pie-chart">${slices}<circle cx="${cx}" cy="${cy}" r="${hole}" fill="var(--panel-bg)" /></svg>`;\n    const center = `<div class="pie-chart-center"><div class="pie-total">${Math.round(total)}</div><div class="pie-total-label">Total FTE</div></div>`;\n    const legend = `<div class="pie-chart-legend">${data.map(d =>\n      `<span class="legend-item"><span class="legend-swatch" style="background:${categoryColor(d.category)}"></span>${d.category}<span class="legend-pct">${((d.fte / total) * 100).toFixed(0)}%</span><span class="legend-value">${d.fte}</span></span>`\n    ).join("")}</div>`;\n    el.innerHTML = `<div class="pie-chart-wrap">${svg}${center}${legend}</div>`;\n  }\n\n  function renderLocationBar(elId, breakdown) {\n    const el = qs(elId);\n    if (!el) return;\n    const categories = breakdown.categories || [];\n    const rows = breakdown.by_location || [];\n    if (!categories.length || !rows.length) {\n      el.innerHTML = \'<div class="chart-empty">No category breakdown for the current filters.</div>\';\n      return;\n    }\n    const maxTotal = Math.max(...rows.map(r => r.total), 1);\n    const barRows = rows.map(r => {\n      const segs = categories.map(cat => {\n        const val = r[cat] || 0;\n        if (val <= 0) return "";\n        const pct = (val / maxTotal) * 100;\n        return `<div class="bar-seg" style="width:${pct}%;background:${categoryColor(cat)}" title="${cat}: ${val}"></div>`;\n      }).join("");\n      return `<div class="bar-row">\n        <div class="bar-label" title="${r.location}">${r.location}</div>\n        <div class="bar-track">${segs}</div>\n        <div class="bar-total">${r.total}</div>\n      </div>`;\n    }).join("");\n    el.innerHTML = `<div class="bar-chart">${barRows}</div>${renderCategoryLegend(categories)}`;\n  }\n\n  // -------------------------------------------------------------------\n  // Generic table renderers\n  // -------------------------------------------------------------------\n  function renderSummaryCards(summary) {\n    const el = qs("#summary-cards");\n    if (!el) return;\n    const cards = [\n      ["Customers", summary.customers],\n      ["Locations", summary.locations],\n      ["Total FTE", summary.total_fte],\n      ["Base Locations", summary.base_locations],\n      ["Additional Locations", summary.additional_locations],\n      ["Transfer Opportunities", summary.transfer_opportunities],\n    ];\n    el.innerHTML = cards.map(([label, val]) =>\n      `<div class="card"><div class="card-label">${label}</div><div class="card-value">${val}</div></div>`\n    ).join("");\n  }\n\n  function renderLocationSummary(rows) {\n    const table = qs("#location-summary-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>No data.</td></tr>"; return; }\n    let html = "<thead><tr><th>Location</th><th>Total FTE</th><th>Customers Supported</th><th>Base For</th></tr></thead><tbody>";\n    rows.forEach(r => {\n      html += `<tr><td>${r.location}</td><td>${r.total_fte}</td><td>${r.customer_count}</td><td>${r.base_for_count}</td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n  }\n\n  function renderCustomerSummary(rows) {\n    const table = qs("#customer-summary-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>No data.</td></tr>"; return; }\n    let html = "<thead><tr><th>Customer</th><th>Base Location(s)</th><th>Additional Locations</th><th>Total FTE</th><th>Status</th></tr></thead><tbody>";\n    rows.forEach(r => {\n      const status = r.warning ? `<span class="status-reduction">⚠ ${r.warning}</span>` : `<span class="status-growth">OK</span>`;\n      html += `<tr><td>${r.customer}</td><td>${r.base_locations.join(", ") || "-"}</td><td>${r.additional_locations.join(", ") || "-"}</td><td>${r.total_fte}</td><td>${status}</td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n  }\n\n  function statusClass(status) {\n    return { Growth: "status-growth", Reduction: "status-reduction", New: "status-new", Discontinued: "status-discontinued", "No Change": "status-stable" }[status] || "";\n  }\n\n  function renderComparison(rows) {\n    const table = qs("#comparison-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>Select two different periods to compare.</td></tr>"; return; }\n    let html = "<thead><tr><th>Customer</th><th>Location</th><th>From FTE</th><th>To FTE</th><th>Change</th><th>Change %</th><th>Status</th></tr></thead><tbody>";\n    rows.forEach(r => {\n      html += `<tr><td>${r.customer}</td><td>${r.location}</td><td>${r.from_fte}</td><td>${r.to_fte}</td><td>${r.change}</td><td>${r.change_pct}%</td><td class="${statusClass(r.status)}">${r.status}</td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n  }\n\n  function renderCapacity(rows) {\n    const table = qs("#capacity-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>Select two different periods to compare.</td></tr>"; return; }\n    let html = "<thead><tr><th>Location</th><th>Current FTE</th><th>Future FTE</th><th>Change</th><th>Status</th></tr></thead><tbody>";\n    rows.forEach(r => {\n      const cls = r.status === "Growth" ? "status-growth" : r.status === "Potential Released Capacity" ? "status-reduction" : "status-stable";\n      html += `<tr><td>${r.location}</td><td>${r.current_fte}</td><td>${r.future_fte}</td><td>${r.change}</td><td class="${cls}">${r.status}</td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n  }\n\n  function renderTransfer(rows) {\n    const table = qs("#transfer-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>No transfer opportunities for the selected periods.</td></tr>"; return; }\n    let html = "<thead><tr><th>From Location (Releasing)</th><th>To Location (Growing)</th><th>Released FTE</th><th>Required FTE</th><th>Matchable FTE</th></tr></thead><tbody>";\n    rows.forEach(r => {\n      html += `<tr><td>${r.from_location}</td><td>${r.to_location}</td><td>${r.released_fte}</td><td>${r.required_fte}</td><td><strong>${r.matchable_fte}</strong></td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n  }\n\n  function renderManualMappings(rows) {\n    const table = qs("#manual-mapping-table");\n    if (!table) return;\n    if (!rows.length) { table.innerHTML = "<tr><td>No manual mappings yet.</td></tr>"; return; }\n    let html = "<thead><tr><th>From</th><th>To</th><th>FTE</th><th>Notes</th><th></th></tr></thead><tbody>";\n    rows.forEach(r => {\n      html += `<tr><td>${r.from_customer} / ${r.from_location}</td><td>${r.to_customer} / ${r.to_location}</td><td>${r.fte_amount}</td><td>${r.notes || ""}</td><td><button data-id="${r.id}" class="btn btn-link delete-mapping">Remove</button></td></tr>`;\n    });\n    table.innerHTML = html + "</tbody>";\n    qsa(".delete-mapping", table).forEach(btn => {\n      btn.addEventListener("click", async () => {\n        await fetch(`/api/manual-mapping/${btn.dataset.id}`, { method: "DELETE", credentials: "same-origin" });\n        loadManualMappings();\n      });\n    });\n  }\n\n  function updateExportLink() {\n    const link = qs("#export-link");\n    if (!link) return;\n    const params = new URLSearchParams();\n    params.set("period", state.period || "");\n    params.set("category", state.category || "Total");\n    if (state.view) params.set("view", state.view);\n    if (state.fromPeriod) params.set("from_period", state.fromPeriod);\n    if (state.toPeriod) params.set("to_period", state.toPeriod);\n    for (const [dim, values] of Object.entries(state.filters)) {\n      values.forEach(v => params.append(dim, v));\n    }\n    link.href = "/export?" + params.toString();\n  }\n\n  // -------------------------------------------------------------------\n  // Page loaders\n  // -------------------------------------------------------------------\n  async function loadDashboardData() {\n    const [summary, matrix, locSummary, custSummary, breakdown] = await Promise.all([\n      api("/api/summary", { period: state.period, category: state.category, ...currentFilterParams() }),\n      api("/api/matrix", { period: state.period, category: state.category, ...currentFilterParams() }),\n      api("/api/location-summary", { period: state.period, category: state.category, ...currentFilterParams() }),\n      api("/api/customer-summary", { period: state.period, category: state.category, ...currentFilterParams() }),\n      api("/api/category-breakdown", { period: state.period, ...currentFilterParams() }),\n    ]);\n    renderSummaryCards(summary);\n    renderMatrix(matrix);\n    renderLocationSummary(locSummary);\n    renderCustomerSummary(custSummary);\n    renderCategoryPie("#category-pie", breakdown);\n    renderLocationBar("#location-bar", breakdown);\n    updateExportLink();\n  }\n\n  async function loadMatrixOnly() {\n    if (state.view === "breakdown") {\n      const matrix = await api("/api/matrix-breakdown", { period: state.period, ...currentFilterParams() });\n      renderMatrixBreakdown(matrix);\n    } else {\n      const matrix = await api("/api/matrix", { period: state.period, category: state.category, ...currentFilterParams() });\n      renderMatrix(matrix);\n      const legendEl = qs("#matrix-category-legend");\n      if (legendEl) legendEl.innerHTML = "";\n    }\n    updateExportLink();\n  }\n\n  async function loadAnalysisData() {\n    const [comparison, capacity] = await Promise.all([\n      api("/api/comparison", { from_period: state.fromPeriod, to_period: state.toPeriod, category: state.category, ...currentFilterParams() }),\n      api("/api/capacity", { from_period: state.fromPeriod, to_period: state.toPeriod, category: state.category, ...currentFilterParams() }),\n    ]);\n    renderComparison(comparison);\n    renderCapacity(capacity);\n    updateExportLink();\n  }\n\n  async function loadTransferData() {\n    const rows = await api("/api/transfer-opportunities", { from_period: state.fromPeriod, to_period: state.toPeriod, category: state.category, ...currentFilterParams() });\n    renderTransfer(rows);\n    updateExportLink();\n  }\n\n  async function loadManualMappings() {\n    const rows = await api("/api/manual-mapping");\n    renderManualMappings(rows);\n  }\n\n  async function loadManagementData() {\n    await loadDashboardData();\n    await loadAnalysisData_ManagementSubset();\n  }\n\n  async function loadAnalysisData_ManagementSubset() {\n    const [comparison, transfers] = await Promise.all([\n      api("/api/comparison", { from_period: state.fromPeriod, to_period: state.toPeriod, category: state.category, ...currentFilterParams() }),\n      api("/api/transfer-opportunities", { from_period: state.fromPeriod, to_period: state.toPeriod, category: state.category, ...currentFilterParams() }),\n    ]);\n    renderComparison(comparison);\n    renderTransfer(transfers.slice(0, 10));\n    updateExportLink();\n  }\n\n  function initDashboard() {\n    buildFilterBar(loadDashboardData);\n    loadDashboardData();\n  }\n\n  function initMapping() {\n    buildFilterBar(loadMatrixOnly);\n    qsa("#view-toggle button").forEach(btn => {\n      btn.addEventListener("click", () => {\n        qsa("#view-toggle button").forEach(b => b.classList.remove("active"));\n        btn.classList.add("active");\n        state.view = btn.dataset.view;\n        loadMatrixOnly();\n      });\n    });\n    loadMatrixOnly();\n  }\n\n  function initAnalysis() {\n    buildFilterBar(loadAnalysisData);\n    loadAnalysisData();\n  }\n\n  function initTransfer() {\n    buildFilterBar(loadTransferData);\n    loadTransferData();\n    loadManualMappings();\n    const form = qs("#manual-mapping-form");\n    form.addEventListener("submit", async (e) => {\n      e.preventDefault();\n      const fd = new FormData(form);\n      await fetch("/api/manual-mapping", { method: "POST", body: fd, credentials: "same-origin" });\n      form.reset();\n      loadManualMappings();\n    });\n  }\n\n  function initManagement() {\n    buildFilterBar(loadManagementData);\n    loadManagementData();\n  }\n\n  // -------------------------------------------------------------------\n  // Assistant widget: a floating chat popup available on every page\n  // (not a separate tab). Quick-question pills, a "Build a Question"\n  // metric picker, and a free-text box that resolves locally to a\n  // predefined question or a customer/location lookup -- no external\n  // AI/LLM call, every answer comes from the same analysis functions\n  // that power the rest of the app.\n  // -------------------------------------------------------------------\n  const widget = { period: null, fromPeriod: null, toPeriod: null };\n\n  function mdBold(text) {\n    const span = document.createElement("span");\n    span.textContent = text;\n    return span.innerHTML.replace(/\\*\\*(.+?)\\*\\*/g, "<strong>$1</strong>");\n  }\n\n  function widgetFeed() { return qs("#assistant-widget-feed"); }\n\n  function clearWidgetHint() {\n    const feed = widgetFeed();\n    const hint = feed && feed.querySelector(".assistant-hint");\n    if (hint) hint.remove();\n  }\n\n  function addChatQuestion(feed, text) {\n    if (!feed) return;\n    clearWidgetHint();\n    const div = document.createElement("div");\n    div.className = "chat-entry chat-question";\n    const bubble = document.createElement("div");\n    bubble.className = "chat-bubble";\n    bubble.textContent = text;\n    div.appendChild(bubble);\n    feed.appendChild(div);\n    feed.scrollTop = feed.scrollHeight;\n  }\n\n  function addChatAnswer(feed, result) {\n    if (!feed) return;\n    const div = document.createElement("div");\n    div.className = "chat-entry chat-answer";\n    let html = `<div class="chat-bubble">${mdBold(result.summary)}</div>`;\n    if (result.table && result.table.rows.length) {\n      html += `<div class="table-scroll"><table class="data-table chat-table">\n        <thead><tr>${result.table.columns.map(c => `<th>${c}</th>`).join("")}</tr></thead>\n        <tbody>${result.table.rows.map(row => `<tr>${row.map(cell => `<td>${cell}</td>`).join("")}</tr>`).join("")}</tbody>\n      </table></div>`;\n    }\n    div.innerHTML = html;\n    feed.appendChild(div);\n    feed.scrollTop = feed.scrollHeight;\n  }\n\n  function addChatError(feed, err) {\n    if (!feed) return;\n    const div = document.createElement("div");\n    div.className = "chat-entry chat-error";\n    const bubble = document.createElement("div");\n    bubble.className = "chat-bubble";\n    bubble.textContent = `⚠ ${err.message}`;\n    div.appendChild(bubble);\n    feed.appendChild(div);\n    feed.scrollTop = feed.scrollHeight;\n  }\n\n  function renderWidgetPills(questions) {\n    const el = qs("#assistant-quick-pills");\n    if (!el) return;\n    el.innerHTML = "";\n\n    const buildBtn = document.createElement("button");\n    buildBtn.type = "button";\n    buildBtn.className = "assistant-pill assistant-pill-accent";\n    buildBtn.textContent = "🧩 Build a Question";\n    buildBtn.addEventListener("click", () => {\n      const builder = qs("#assistant-builder");\n      if (builder) builder.hidden = !builder.hidden;\n    });\n    el.appendChild(buildBtn);\n\n    const shortcuts = [\n      ["🏢", "top_location", "Top Location"],\n      ["👥", "top_customer", "Top Customer"],\n      ["⚠️", "base_issues", "Base Issues"],\n      ["🔻", "releasing_capacity", "Releasing Capacity"],\n      ["📈", "growing_locations", "Growing"],\n      ["🔄", "transfer_opportunities", "Transfers"],\n      ["🥧", "category_split", "Category Split"],\n    ];\n    shortcuts.forEach(([icon, id, label]) => {\n      const q = questions.find(item => item.id === id);\n      if (!q) return;\n      const btn = document.createElement("button");\n      btn.type = "button";\n      btn.className = "assistant-pill";\n      btn.textContent = `${icon} ${label}`;\n      btn.addEventListener("click", () => askWidgetPredefined(q));\n      el.appendChild(btn);\n    });\n  }\n\n  function renderWidgetBuilder(metrics, periods) {\n    const metricSel = qs("#assistant-metric-select");\n    const periodSel = qs("#assistant-period");\n    const fromSel = qs("#assistant-from-period");\n    const toSel = qs("#assistant-to-period");\n    if (!metricSel || !periodSel || !fromSel || !toSel) return;\n\n    metricSel.innerHTML = metrics.map(m => `<option value="${m.id}">${m.label}</option>`).join("");\n    [periodSel, fromSel, toSel].forEach(sel => {\n      sel.innerHTML = periods.map(p => `<option value="${p}">${p}</option>`).join("");\n    });\n    periodSel.value = widget.period;\n    fromSel.value = widget.fromPeriod;\n    toSel.value = widget.toPeriod;\n    periodSel.addEventListener("change", () => { widget.period = periodSel.value; });\n    fromSel.addEventListener("change", () => { widget.fromPeriod = fromSel.value; });\n    toSel.addEventListener("change", () => { widget.toPeriod = toSel.value; });\n\n    qs("#assistant-builder-ask").addEventListener("click", async () => {\n      const opt = metricSel.options[metricSel.selectedIndex];\n      const feed = widgetFeed();\n      addChatQuestion(feed, opt.textContent);\n      try {\n        const result = await api("/api/assistant/query", {\n          metric_id: metricSel.value, period: widget.period, from_period: widget.fromPeriod, to_period: widget.toPeriod,\n        });\n        addChatAnswer(feed, result);\n      } catch (err) {\n        addChatError(feed, err);\n      }\n    });\n  }\n\n  async function askWidgetPredefined(q) {\n    const feed = widgetFeed();\n    addChatQuestion(feed, q.label);\n    try {\n      const result = await api("/api/assistant/ask", {\n        question_id: q.id, period: widget.period, from_period: widget.fromPeriod, to_period: widget.toPeriod,\n      });\n      addChatAnswer(feed, result);\n    } catch (err) {\n      addChatError(feed, err);\n    }\n  }\n\n  async function askWidgetFreeform(text) {\n    const feed = widgetFeed();\n    addChatQuestion(feed, text);\n    try {\n      const result = await api("/api/assistant/freeform", {\n        q: text, period: widget.period, from_period: widget.fromPeriod, to_period: widget.toPeriod,\n      });\n      addChatAnswer(feed, result);\n    } catch (err) {\n      addChatError(feed, err);\n    }\n  }\n\n  function toggleWidgetPanel(show) {\n    const panel = qs("#assistant-panel");\n    if (panel) panel.hidden = !show;\n  }\n\n  async function initAssistantWidget() {\n    const root = qs("#assistant-widget");\n    if (!root) return;\n\n    let meta, assistantMeta;\n    try {\n      meta = await api("/api/meta");\n      assistantMeta = await api("/api/assistant/meta");\n    } catch (_) {\n      root.hidden = true; // no dataset loaded on this page yet\n      return;\n    }\n\n    widget.period = meta.periods[meta.periods.length - 1] || null;\n    widget.fromPeriod = meta.periods[0] || null;\n    widget.toPeriod = meta.periods[meta.periods.length - 1] || null;\n\n    root.hidden = false;\n    renderWidgetPills(assistantMeta.questions);\n    renderWidgetBuilder(assistantMeta.metrics, meta.periods);\n\n    qs("#assistant-fab").addEventListener("click", () => toggleWidgetPanel(true));\n    qs("#assistant-close").addEventListener("click", () => toggleWidgetPanel(false));\n    qs("#assistant-input-form").addEventListener("submit", (e) => {\n      e.preventDefault();\n      const input = qs("#assistant-input");\n      const text = input.value.trim();\n      if (!text) return;\n      input.value = "";\n      askWidgetFreeform(text);\n    });\n  }\n\n  return {\n    initDashboard, initMapping, initAnalysis, initTransfer, initManagement, initAssistantWidget,\n  };\n})();\n'
class SimpleTemplates:
    """Minimal stand-in for Starlette's Jinja2Templates, backed by an
    in-memory DictLoader so no templates/ directory is needed on disk."""

    def __init__(self, templates: dict):
        self.env = Environment(loader=DictLoader(templates), autoescape=True)
        self.env.filters["tojson"] = lambda obj, **kw: json.dumps(obj)

    def TemplateResponse(self, name, context, status_code: int = 200):
        html = self.env.get_template(name).render(**context)
        return HTMLResponse(html, status_code=status_code)

# ============================================================================
# --- from app/models.py ---
# ============================================================================
"""
Dataclasses / typed containers used throughout the FTE Location Planner.

The application never hard-codes business values (customer names, location
names, periods, etc). Everything here is a generic container that is filled
in dynamically from whatever Excel file the user uploads.
"""

from dataclasses import dataclass, field
from typing import Optional
import pandas as pd


# Canonical FTE categories the app understands. "Total" is always included
# because every uploaded workbook has a period-level total column
# (e.g. FTE_Dec_2025). Category-level columns (Internal/SWC/External/Others)
# are optional -- if a workbook doesn't have them, only "Total" is offered.
CANONICAL_CATEGORIES = ["Internal", "SWC", "External", "Others"]
TOTAL_CATEGORY = "Total"


@dataclass
class DetectedColumns:
    """Result of scanning the uploaded workbook's header row."""
    sno: Optional[str] = None
    vm_product: Optional[str] = None
    customer_account: Optional[str] = None
    component: Optional[str] = None
    country: Optional[str] = None
    location: Optional[str] = None
    base_location: Optional[str] = None
    # period -> {category_or_'Total': column_name}
    fte_columns: dict = field(default_factory=dict)
    unmatched_columns: list = field(default_factory=list)
    missing_required: list = field(default_factory=list)


@dataclass
class Dataset:
    """
    Fully parsed & normalized representation of one uploaded workbook.

    raw_df   : the original dataframe exactly as uploaded (for export/reference)
    long_df  : normalized long-format data, one row per
               (S.No, dimension columns..., Period, Category, FTE)
    dims     : dict of dimension name -> sorted list of distinct values found
    periods  : ordered list of period labels found in the workbook, in
               chronological order as they appeared left-to-right in Excel
    categories: list of FTE categories available (subset of CANONICAL_CATEGORIES + Total)
    warnings : data-quality warnings collected while parsing
    """
    raw_df: pd.DataFrame
    long_df: pd.DataFrame
    dims: dict
    periods: list
    categories: list
    warnings: list
    detected_columns: DetectedColumns
    source_filename: str

# ============================================================================
# --- from app/utils.py ---
# ============================================================================
"""
Generic helpers shared across the app:
- flexible column-name matching (so slightly different header spellings
  from different Excel files still work)
- Base location value normalization (Yes/No/Y/N/TRUE/FALSE/1/0)
- a tiny in-memory store so a user's parsed dataset survives between
  page navigations without needing a database
"""

import re
import threading
import uuid
from typing import Optional

# ---------------------------------------------------------------------------
# Column name matching
# ---------------------------------------------------------------------------

def _normalize_header(name: str) -> str:
    """Lowercase, strip punctuation/whitespace so headers compare loosely."""
    name = str(name).strip().lower()
    name = re.sub(r"[^a-z0-9]+", "", name)
    return name


# Candidate spellings for each required dimension column. New synonyms can
# be added here without touching any other part of the app.
COLUMN_SYNONYMS = {
    "sno": ["sno", "s.no", "serialno", "slno", "sl.no", "id", "rowno"],
    "vm_product": ["vmproduct", "product", "vm"],
    "customer_account": ["customeraccount", "customer", "account", "client"],
    "component": ["component", "workstream", "tower"],
    "country": ["country"],
    "location": ["location", "site", "city"],
    "base_location": ["baselocation", "isbase", "basesite", "base"],
}


def find_column(actual_columns: list[str], field_key: str) -> Optional[str]:
    """Return the actual column name in the workbook that best matches
    the given logical field (e.g. 'customer_account')."""
    synonyms = {_normalize_header(s) for s in COLUMN_SYNONYMS[field_key]}
    for col in actual_columns:
        if _normalize_header(col) in synonyms:
            return col
    return None


# FTE columns look like FTE_<Month>_<Year> or FTE_<Month>_<Year>_<Category>
FTE_COLUMN_RE = re.compile(
    r"^FTE[_\s]*([A-Za-z]{3,9})[_\s]*(\d{4})(?:[_\s]+([A-Za-z]+))?$",
    re.IGNORECASE,
)

MONTH_ORDER = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}

CATEGORY_ALIASES = {
    "internal": "Internal",
    "swc": "SWC",
    "external": "External",
    "others": "Others",
    "other": "Others",
}


def parse_fte_column(col_name: str):
    """
    Given a column name, return (period_label, category) if it looks like
    an FTE column, else None.

    period_label is normalized to 'Mon-YYYY' (e.g. 'Dec-2025') and used
    for both display and chronological sorting.
    category is one of Internal/SWC/External/Others/Total.
    """
    m = FTE_COLUMN_RE.match(str(col_name).strip())
    if not m:
        return None
    month_raw, year, cat_raw = m.groups()
    month_key = month_raw.lower()
    if month_key not in MONTH_ORDER:
        return None
    month_num = MONTH_ORDER[month_key]
    month_abbrev = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    ][month_num - 1]
    period_label = f"{month_abbrev}-{year}"
    sort_key = int(year) * 100 + month_num

    if cat_raw is None:
        category = "Total"
    else:
        category = CATEGORY_ALIASES.get(cat_raw.lower())
        if category is None:
            return None
    return period_label, category, sort_key


# ---------------------------------------------------------------------------
# Base location value normalization
# ---------------------------------------------------------------------------

TRUE_VALUES = {"yes", "y", "true", "1", "1.0"}
FALSE_VALUES = {"no", "n", "false", "0", "0.0", "", "none", "nan"}


def normalize_base_flag(value) -> Optional[bool]:
    """Return True/False, or None if the value is blank/unrecognized."""
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return None


# ---------------------------------------------------------------------------
# In-memory dataset store (single-process; fine for this tool's scale)
# ---------------------------------------------------------------------------

class DatasetStore:
    def __init__(self):
        self._lock = threading.Lock()
        self._store: dict[str, "Dataset"] = {}

    def new_id(self) -> str:
        return uuid.uuid4().hex

    def put(self, dataset_id: str, dataset) -> None:
        with self._lock:
            self._store[dataset_id] = dataset

    def get(self, dataset_id: str):
        with self._lock:
            return self._store.get(dataset_id)

    def delete(self, dataset_id: str) -> None:
        with self._lock:
            self._store.pop(dataset_id, None)


store = DatasetStore()
SESSION_COOKIE = "fte_dataset_id"

# ============================================================================
# --- from app/excel_processor.py ---
# ============================================================================
"""
Reads an uploaded Excel workbook and converts it into a normalized,
long-format dataset, regardless of exact column order/naming, number of
rows, number of locations, number of periods, etc.

This is the only place that touches raw Excel structure. Everything
downstream (mapping_engine, analysis_engine) works purely off the
normalized Dataset object.
"""

import logging
from pathlib import Path

import pandas as pd


excel_logger = logging.getLogger("fte_planner.excel_processor")

REQUIRED_FIELDS = ["customer_account", "location"]


class ExcelProcessingError(Exception):
    """Raised for any problem that should surface as a friendly error page."""


# Some workbooks (e.g. the standard "FTE_Location_Planner" template) use a
# two-row header: row 1 has the period label merged across a group of
# columns (e.g. "FTE_Dec_2025" spanning 5 columns), row 2 has the
# per-column breakdown ("Internal", "SWC", "External", "Others", and the
# period's own total repeated). Plain pandas can't read that directly, so
# it's flattened here into single-row headers like "FTE_Dec_2025_Internal"
# that `parse_fte_column` already understands.
CATEGORY_HEADER_ALIASES = {"internal", "swc", "external", "others", "other"}


def _looks_like_two_row_header(raw: pd.DataFrame) -> bool:
    """True if row 2 (index 1) looks like Internal/SWC/External/Others sub-headers."""
    if len(raw) < 3:
        return False
    row2 = raw.iloc[1]
    matches = sum(
        1 for v in row2
        if isinstance(v, str) and v.strip().lower() in CATEGORY_HEADER_ALIASES
    )
    return matches >= 2


def _flatten_two_row_header(raw: pd.DataFrame) -> pd.DataFrame:
    row1 = list(raw.iloc[0])
    row2 = list(raw.iloc[1])

    # Excel merged cells only store a value in the top-left cell of the
    # range; the rest read back as NaN. Forward-fill row 1 so every column
    # in a merged period group picks up that group's label.
    filled_row1 = []
    last = None
    for v in row1:
        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "":
            filled_row1.append(last)
        else:
            last = v
            filled_row1.append(v)

    combined = []
    for top, sub in zip(filled_row1, row2):
        if sub is not None and not (isinstance(sub, float) and pd.isna(sub)) and str(sub).strip() != "":
            sub_text = str(sub).strip()
            if sub_text.lower() in CATEGORY_HEADER_ALIASES:
                combined.append(f"{top}_{sub_text}")
            else:
                # The group's total sub-header repeats the period label
                # itself (e.g. "FTE_Dec_2025") -- use it as-is.
                combined.append(sub_text)
        else:
            combined.append(top)

    data = raw.iloc[2:].reset_index(drop=True)
    data.columns = [str(c).strip() if c is not None else "" for c in combined]
    return data


def _read_sheet(xls_or_path, sheet_name=None, engine: str | None = None) -> pd.DataFrame:
    kwargs = {"header": None}
    if sheet_name is not None:
        kwargs["sheet_name"] = sheet_name
    if engine is not None:
        kwargs["engine"] = engine
    raw = pd.read_excel(xls_or_path, **kwargs)
    if _looks_like_two_row_header(raw):
        return _flatten_two_row_header(raw)

    header_kwargs = {k: v for k, v in kwargs.items() if k != "header"}
    return pd.read_excel(xls_or_path, **header_kwargs)


def _read_workbook(path: Path) -> pd.DataFrame:
    try:
        if path.suffix.lower() == ".xls":
            df = _read_sheet(path, engine="xlrd")
        else:
            # Pick the first sheet that actually looks like tabular data
            # (more than 1 column, header row present). The `with` here
            # matters on Windows: an unclosed ExcelFile keeps the temp
            # upload locked, so the caller's later os.unlink() fails.
            with pd.ExcelFile(path) as xls:
                best_sheet = None
                best_cols = -1
                for name in xls.sheet_names:
                    try:
                        preview = pd.read_excel(xls, sheet_name=name, nrows=5)
                    except Exception:
                        continue
                    if preview.shape[1] > best_cols and preview.shape[1] > 2:
                        best_cols = preview.shape[1]
                        best_sheet = name
                if best_sheet is None:
                    best_sheet = xls.sheet_names[0]
                df = _read_sheet(xls, sheet_name=best_sheet)
    except Exception as exc:  # noqa: BLE001
        excel_logger.exception("Failed reading workbook")
        raise ExcelProcessingError("File is corrupted or not a valid Excel file") from exc

    if df.empty or df.shape[1] < 2:
        raise ExcelProcessingError("Excel format is invalid or file has no data")
    return df


def detect_columns(df: pd.DataFrame) -> DetectedColumns:
    cols = list(df.columns)
    detected = DetectedColumns(
        sno=find_column(cols, "sno"),
        vm_product=find_column(cols, "vm_product"),
        customer_account=find_column(cols, "customer_account"),
        component=find_column(cols, "component"),
        country=find_column(cols, "country"),
        location=find_column(cols, "location"),
        base_location=find_column(cols, "base_location"),
    )

    consumed = {
        v for v in [
            detected.sno, detected.vm_product, detected.customer_account,
            detected.component, detected.country, detected.location,
            detected.base_location,
        ] if v
    }

    fte_columns: dict[str, dict[str, str]] = {}
    unmatched = []
    for col in cols:
        if col in consumed:
            continue
        parsed = parse_fte_column(col)
        if parsed is None:
            unmatched.append(col)
            continue
        period_label, category, sort_key = parsed
        bucket = fte_columns.setdefault(period_label, {"__sort_key__": sort_key})
        bucket[category] = col

    detected.fte_columns = fte_columns
    detected.unmatched_columns = unmatched

    missing = [f for f in REQUIRED_FIELDS if getattr(detected, f) is None]
    if not fte_columns:
        missing.append("fte_columns")
    detected.missing_required = missing
    return detected


def _melt_to_long(df: pd.DataFrame, detected: DetectedColumns) -> pd.DataFrame:
    dim_map = {
        "SNo": detected.sno,
        "VM_Product": detected.vm_product,
        "Customer_Account": detected.customer_account,
        "Component": detected.component,
        "Country": detected.country,
        "Location": detected.location,
    }
    dim_cols = {k: v for k, v in dim_map.items() if v}

    work = pd.DataFrame(index=df.index)
    for logical_name, actual_col in dim_cols.items():
        work[logical_name] = df[actual_col].astype(str).str.strip()

    if detected.base_location:
        work["Base_Location_Flag"] = df[detected.base_location].apply(normalize_base_flag)
    else:
        work["Base_Location_Flag"] = None

    work["_row_id"] = df.index

    periods_sorted = sorted(
        detected.fte_columns.items(), key=lambda kv: kv[1]["__sort_key__"]
    )

    records = []
    for period_label, cat_cols in periods_sorted:
        for category, col_name in cat_cols.items():
            if category == "__sort_key__":
                continue
            values = pd.to_numeric(df[col_name], errors="coerce")
            for idx in df.index:
                fte_val = values.loc[idx]
                if pd.isna(fte_val):
                    continue
                records.append((idx, period_label, category, float(fte_val)))

    long_records = pd.DataFrame(
        records, columns=["_row_id", "Period", "Category", "FTE"]
    )
    long_df = long_records.merge(work, on="_row_id", how="left")
    return long_df


def process_workbook(path: Path, original_filename: str) -> Dataset:
    df = _read_workbook(path)
    # Drop fully-empty rows/columns that sometimes trail an Excel export
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    detected = detect_columns(df)
    if detected.missing_required:
        pretty = ", ".join(detected.missing_required)
        raise ExcelProcessingError(f"Required column(s) could not be detected: {pretty}")

    long_df = _melt_to_long(df, detected)

    warnings = []
    if not detected.base_location:
        warnings.append(
            "No 'Base location' column was found — base-location markers cannot be shown."
        )
    if detected.unmatched_columns:
        warnings.append(
            f"{len(detected.unmatched_columns)} column(s) were not recognized and were ignored: "
            + ", ".join(str(c) for c in detected.unmatched_columns[:10])
            + ("..." if len(detected.unmatched_columns) > 10 else "")
        )

    periods_sorted = [
        p for p, _ in sorted(detected.fte_columns.items(), key=lambda kv: kv[1]["__sort_key__"])
    ]
    categories_present = set()
    for cat_cols in detected.fte_columns.values():
        categories_present.update(k for k in cat_cols if k != "__sort_key__")
    ordered_categories = [c for c in ["Internal", "SWC", "External", "Others"] if c in categories_present]
    if "Total" in categories_present:
        ordered_categories.append("Total")
    elif ordered_categories:
        # No explicit total column — we'll compute Total as a derived sum downstream.
        ordered_categories.append("Total")

    dims = {}
    for dim in ["VM_Product", "Customer_Account", "Component", "Country", "Location"]:
        if dim in long_df.columns:
            dims[dim] = sorted(x for x in long_df[dim].dropna().unique().tolist() if x and x != "nan")

    negative_rows = long_df[long_df["FTE"] < 0]
    if not negative_rows.empty:
        warnings.append(
            f"{len(negative_rows)} record(s) contain negative FTE values — flagged for review."
        )

    return Dataset(
        raw_df=df,
        long_df=long_df,
        dims=dims,
        periods=periods_sorted,
        categories=ordered_categories,
        warnings=warnings,
        detected_columns=detected,
        source_filename=original_filename,
    )

# ============================================================================
# --- from app/mapping_engine.py ---
# ============================================================================
"""
Builds the core "Location Mapping Matrix": Customer Account (rows) x
Location (columns), with base-location markers, supporting-location
checkmarks, and FTE values for whichever period/category is selected.
"""

import pandas as pd


FILTERABLE_DIMS = ["VM_Product", "Customer_Account", "Component", "Country", "Location"]


def apply_filters(long_df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """filters: dict of dim_name -> list[str] (empty/missing list == All)."""
    df = long_df
    for dim in FILTERABLE_DIMS:
        values = filters.get(dim)
        if values:
            df = df[df[dim].isin(values)]
    return df


def fte_for_period_category(long_df: pd.DataFrame, period: str, category: str) -> pd.DataFrame:
    """Return rows for one period, resolving 'Total' whether or not the
    workbook had an explicit Total column."""
    df = long_df[long_df["Period"] == period]
    if category == "Total":
        if (df["Category"] == "Total").any():
            return df[df["Category"] == "Total"]
        return df[df["Category"] != "Total"]
    return df[df["Category"] == category]


def base_location_status(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    Per Customer_Account x Location, whether it is ever flagged as a base
    location (independent of period/category — base location is a
    structural attribute of the account/product/component/location combo).
    Returns columns: Customer_Account, Location, is_base
    """
    dedup = long_df.drop_duplicates(subset=["_row_id"])
    grouped = (
        dedup.groupby(["Customer_Account", "Location"])["Base_Location_Flag"]
        .apply(lambda s: bool(s.fillna(False).any()))
        .reset_index()
        .rename(columns={"Base_Location_Flag": "is_base"})
    )
    return grouped


def customer_base_warnings(base_status: pd.DataFrame) -> dict:
    """Customer_Account -> warning string, or None if exactly one base."""
    warnings = {}
    for customer, grp in base_status.groupby("Customer_Account"):
        base_count = int(grp["is_base"].sum())
        if base_count == 0:
            warnings[customer] = "BASE LOCATION NOT DEFINED"
        elif base_count > 1:
            warnings[customer] = "MULTIPLE BASE LOCATIONS"
        else:
            warnings[customer] = None
    return warnings


def build_matrix(long_df: pd.DataFrame, filters: dict, period: str, category: str) -> dict:
    """
    Returns:
    {
      "customers": [...],
      "locations": [...],
      "cells": {customer: {location: {"fte": float|None, "is_base": bool, "has_fte": bool}}},
      "customer_warnings": {customer: str|None},
    }
    """
    filtered = apply_filters(long_df, filters)

    base_status = base_location_status(filtered)
    warnings = customer_base_warnings(base_status)

    period_df = fte_for_period_category(filtered, period, category) if period else filtered.iloc[0:0]
    fte_grouped = (
        period_df.groupby(["Customer_Account", "Location"])["FTE"].sum().reset_index()
    )

    customers = sorted(filtered["Customer_Account"].dropna().unique().tolist())
    locations = sorted(filtered["Location"].dropna().unique().tolist())

    base_lookup = {(r.Customer_Account, r.Location): r.is_base for r in base_status.itertuples()}
    fte_lookup = {(r.Customer_Account, r.Location): r.FTE for r in fte_grouped.itertuples()}

    cells: dict = {}
    for cust in customers:
        cells[cust] = {}
        for loc in locations:
            is_base = base_lookup.get((cust, loc), False)
            fte_val = fte_lookup.get((cust, loc))
            has_fte = fte_val is not None and fte_val != 0
            cells[cust][loc] = {
                "fte": fte_val,
                "is_base": is_base,
                "has_fte": has_fte,
            }

    return {
        "customers": customers,
        "locations": locations,
        "cells": cells,
        "customer_warnings": warnings,
    }


def build_matrix_breakdown(long_df: pd.DataFrame, filters: dict, period: str) -> dict:
    """
    Like build_matrix, but instead of one FTE number per cell, returns the
    full Internal/SWC/External/Others composition (whichever of those the
    workbook actually has) so the UI can render e.g. "14 + 1 + 3 + 0 = 18"
    with each number color-coded by category.

    Returns:
    {
      "customers": [...],
      "locations": [...],
      "categories": [...],  # subset of CANONICAL_CATEGORIES present in the data
      "cells": {customer: {location: {"values": {cat: fte}, "total": float,
                                       "is_base": bool, "has_fte": bool}}},
      "customer_warnings": {customer: str|None},
    }
    """
    filtered = apply_filters(long_df, filters)

    base_status = base_location_status(filtered)
    warnings = customer_base_warnings(base_status)

    period_df = filtered[filtered["Period"] == period] if period else filtered.iloc[0:0]
    cat_df = period_df[period_df["Category"].isin(CANONICAL_CATEGORIES)]
    present_categories = [c for c in CANONICAL_CATEGORIES if c in set(cat_df["Category"])]

    grouped = (
        cat_df.groupby(["Customer_Account", "Location", "Category"])["FTE"].sum()
    )

    customers = sorted(filtered["Customer_Account"].dropna().unique().tolist())
    locations = sorted(filtered["Location"].dropna().unique().tolist())
    base_lookup = {(r.Customer_Account, r.Location): r.is_base for r in base_status.itertuples()}

    cells: dict = {}
    for cust in customers:
        cells[cust] = {}
        for loc in locations:
            values = {}
            for cat in present_categories:
                try:
                    values[cat] = float(grouped.loc[(cust, loc, cat)])
                except KeyError:
                    values[cat] = 0.0
            total = sum(values.values())
            cells[cust][loc] = {
                "values": values,
                "total": total,
                "is_base": base_lookup.get((cust, loc), False),
                "has_fte": total != 0,
            }

    return {
        "customers": customers,
        "locations": locations,
        "categories": present_categories,
        "cells": cells,
        "customer_warnings": warnings,
    }

# ============================================================================
# --- from app/analysis_engine.py ---
# ============================================================================
"""
Higher-level analytics built on top of the normalized dataset:
summary cards, location/customer summaries, period-over-period comparison,
capacity classification, and simple transfer-opportunity matching.
"""

import pandas as pd


REDUCTION_THRESHOLD = 0  # any negative change flags "released capacity"


def summary_cards(long_df: pd.DataFrame, filters: dict, period: str, category: str) -> dict:
    filtered = apply_filters(long_df, filters)
    period_df = fte_for_period_category(filtered, period, category)

    base_status = base_location_status(filtered)
    base_counts = base_status.groupby("Customer_Account")["is_base"].sum()

    total_fte = float(period_df["FTE"].sum())
    total_customers = filtered["Customer_Account"].nunique()
    total_locations = filtered["Location"].nunique()
    base_locations = int((base_status["is_base"]).sum())
    additional_locations = int(len(base_status) - base_locations)
    transfer_opportunities = len(identify_transfer_opportunities(long_df, filters, period, period))

    return {
        "customers": total_customers,
        "locations": total_locations,
        "total_fte": round(total_fte, 1),
        "base_locations": base_locations,
        "additional_locations": additional_locations,
        "transfer_opportunities": transfer_opportunities,
        "accounts_with_base_issues": int(sum(1 for v in base_counts if v != 1)),
    }


def location_summary(long_df: pd.DataFrame, filters: dict, period: str, category: str) -> list[dict]:
    filtered = apply_filters(long_df, filters)
    period_df = fte_for_period_category(filtered, period, category)

    grouped = period_df.groupby("Location").agg(
        total_fte=("FTE", "sum"),
        customers=("Customer_Account", lambda s: sorted(set(s))),
    ).reset_index()

    base_status = base_location_status(filtered)
    base_counts = base_status[base_status["is_base"]].groupby("Location").size()

    rows = []
    for r in grouped.itertuples():
        rows.append({
            "location": r.Location,
            "total_fte": round(float(r.total_fte), 1),
            "customer_count": len(r.customers),
            "customers": r.customers,
            "base_for_count": int(base_counts.get(r.Location, 0)),
        })
    rows.sort(key=lambda x: x["total_fte"], reverse=True)
    return rows


def customer_summary(long_df: pd.DataFrame, filters: dict, period: str, category: str) -> list[dict]:
    filtered = apply_filters(long_df, filters)
    period_df = fte_for_period_category(filtered, period, category)
    base_status = base_location_status(filtered)

    grouped = period_df.groupby("Customer_Account")["FTE"].sum()

    rows = []
    for customer, base_grp in base_status.groupby("Customer_Account"):
        base_locs = sorted(base_grp[base_grp["is_base"]]["Location"].tolist())
        additional_locs = sorted(base_grp[~base_grp["is_base"]]["Location"].tolist())
        rows.append({
            "customer": customer,
            "base_locations": base_locs,
            "additional_locations": additional_locs,
            "total_fte": round(float(grouped.get(customer, 0.0)), 1),
            "warning": (
                "MULTIPLE BASE LOCATIONS" if len(base_locs) > 1
                else "BASE LOCATION NOT DEFINED" if len(base_locs) == 0
                else None
            ),
        })
    rows.sort(key=lambda x: x["total_fte"], reverse=True)
    return rows


def category_breakdown(long_df: pd.DataFrame, filters: dict, period: str) -> dict:
    """
    Internal/SWC/External/Others composition for the selected period +
    dimension filters (the "FTE Type" filter itself is ignored here since
    the whole point is to show every category at once) -- feeds the
    Dashboard's pie chart (overall split) and stacked bar chart (by
    location).
    """
    filtered = apply_filters(long_df, filters)
    period_df = filtered[filtered["Period"] == period] if period else filtered.iloc[0:0]
    cat_df = period_df[period_df["Category"].isin(CANONICAL_CATEGORIES)]
    present_categories = [c for c in CANONICAL_CATEGORIES if c in set(cat_df["Category"])]

    by_category = []
    if present_categories:
        totals = cat_df.groupby("Category")["FTE"].sum()
        for cat in present_categories:
            by_category.append({"category": cat, "fte": round(float(totals.get(cat, 0.0)), 1)})

    by_location = []
    if present_categories:
        grouped = cat_df.groupby(["Location", "Category"])["FTE"].sum().unstack(fill_value=0.0)
        grouped = grouped.reindex(columns=present_categories, fill_value=0.0)
        grouped["__total__"] = grouped.sum(axis=1)
        grouped = grouped.sort_values("__total__", ascending=False)
        for loc, row in grouped.iterrows():
            entry = {"location": loc, "total": round(float(row["__total__"]), 1)}
            for cat in present_categories:
                entry[cat] = round(float(row[cat]), 1)
            by_location.append(entry)

    return {
        "categories": present_categories,
        "by_category": by_category,
        "by_location": by_location,
    }


def period_comparison(long_df: pd.DataFrame, filters: dict, from_period: str, to_period: str, category: str) -> list[dict]:
    filtered = apply_filters(long_df, filters)
    from_df = fte_for_period_category(filtered, from_period, category)
    to_df = fte_for_period_category(filtered, to_period, category)

    from_grp = from_df.groupby(["Customer_Account", "Location"])["FTE"].sum()
    to_grp = to_df.groupby(["Customer_Account", "Location"])["FTE"].sum()

    keys = set(from_grp.index) | set(to_grp.index)
    rows = []
    for key in keys:
        customer, location = key
        from_val = float(from_grp.get(key, 0.0))
        to_val = float(to_grp.get(key, 0.0))
        change = to_val - from_val
        pct = (change / from_val * 100) if from_val else (100.0 if to_val else 0.0)
        if from_val == 0 and to_val > 0:
            status = "New"
        elif to_val == 0 and from_val > 0:
            status = "Discontinued"
        elif change > 0:
            status = "Growth"
        elif change < 0:
            status = "Reduction"
        else:
            status = "No Change"
        rows.append({
            "customer": customer,
            "location": location,
            "from_fte": round(from_val, 1),
            "to_fte": round(to_val, 1),
            "change": round(change, 1),
            "change_pct": round(pct, 1),
            "status": status,
        })
    rows.sort(key=lambda x: x["change"])
    return rows


def capacity_view(long_df: pd.DataFrame, filters: dict, from_period: str, to_period: str, category: str) -> list[dict]:
    filtered = apply_filters(long_df, filters)
    from_df = fte_for_period_category(filtered, from_period, category)
    to_df = fte_for_period_category(filtered, to_period, category)

    from_grp = from_df.groupby("Location")["FTE"].sum()
    to_grp = to_df.groupby("Location")["FTE"].sum()
    locations = sorted(set(from_grp.index) | set(to_grp.index))

    rows = []
    for loc in locations:
        current = float(from_grp.get(loc, 0.0))
        future = float(to_grp.get(loc, 0.0))
        change = future - current
        if change < REDUCTION_THRESHOLD:
            status = "Potential Released Capacity"
        elif change > 0:
            status = "Growth"
        else:
            status = "Stable"
        rows.append({
            "location": loc,
            "current_fte": round(current, 1),
            "future_fte": round(future, 1),
            "change": round(change, 1),
            "status": status,
        })
    rows.sort(key=lambda x: x["change"])
    return rows


def identify_transfer_opportunities(long_df: pd.DataFrame, filters: dict, from_period: str, to_period: str, category: str = "Total") -> list[dict]:
    """
    Very simple heuristic: locations releasing FTE (reducing) are matched
    against locations/customers requiring FTE (growing) in the same window,
    ranked by the smaller of the two magnitudes (the "matchable" amount).
    """
    if from_period == to_period:
        return []
    cap = capacity_view(long_df, filters, from_period, to_period, category)
    releasing = [c for c in cap if c["status"] == "Potential Released Capacity"]
    growing = [c for c in cap if c["status"] == "Growth"]

    opportunities = []
    for r in releasing:
        for g in growing:
            matchable = min(abs(r["change"]), abs(g["change"]))
            if matchable <= 0:
                continue
            opportunities.append({
                "from_location": r["location"],
                "to_location": g["location"],
                "released_fte": abs(r["change"]),
                "required_fte": g["change"],
                "matchable_fte": round(matchable, 1),
            })
    opportunities.sort(key=lambda x: x["matchable_fte"], reverse=True)
    return opportunities


def data_quality_report(long_df: pd.DataFrame, dataset_warnings: list[str]) -> dict:
    base_status = base_location_status(long_df)
    no_base = base_status.groupby("Customer_Account")["is_base"].sum()
    issues_missing = sorted(no_base[no_base == 0].index.tolist())
    issues_multi = sorted(no_base[no_base > 1].index.tolist())

    negative = long_df[long_df["FTE"] < 0]
    negative_records = negative[["Customer_Account", "Location", "Period", "Category", "FTE"]].drop_duplicates()

    return {
        "parse_warnings": dataset_warnings,
        "missing_base_customers": issues_missing,
        "multiple_base_customers": issues_multi,
        "negative_fte_records": negative_records.to_dict(orient="records"),
    }

# ============================================================================
# --- from app/export_engine.py ---
# ============================================================================
"""
Builds a multi-sheet analyzed Excel workbook the user can download,
covering the matrix, summaries, comparison, capacity view, transfer
opportunities, manual mappings, and data-quality findings.
"""

import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BASE_FILL = PatternFill("solid", fgColor="FFE699")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_export_workbook(dataset, filters: dict, period: str, category: str,
                           from_period: str, to_period: str,
                           manual_mappings: list[dict], view: str | None = None) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # --- Matrix sheet ---
        # Mirrors whichever matrix view is active on screen: the normal
        # BASE/SUPPORT + Total matrix, or -- when the "Internal / SWC /
        # External Breakdown" view is selected -- a tidy per-category table
        # instead, so the download always matches what's on screen.
        if view == "breakdown":
            breakdown = build_matrix_breakdown(dataset.long_df, filters, period)
            rows = []
            for cust in breakdown["customers"]:
                for loc in breakdown["locations"]:
                    cell = breakdown["cells"][cust][loc]
                    if not cell["has_fte"]:
                        continue
                    row = {
                        "Customer Account": cust,
                        "Location": loc,
                        "Status": "BASE" if cell["is_base"] else "SUPPORT",
                    }
                    for cat in breakdown["categories"]:
                        row[cat] = cell["values"].get(cat, 0.0)
                    row["Total"] = cell["total"]
                    row["Data Quality"] = breakdown["customer_warnings"].get(cust) or "OK"
                    rows.append(row)
            pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=f"Matrix_Breakdown_{period}"[:31])
        else:
            matrix = build_matrix(dataset.long_df, filters, period, category)
            matrix_rows = []
            for cust in matrix["customers"]:
                row = {"Customer Account": cust}
                for loc in matrix["locations"]:
                    cell = matrix["cells"][cust][loc]
                    if cell["is_base"]:
                        label = "BASE"
                    elif cell["has_fte"]:
                        label = "SUPPORT"
                    else:
                        label = "-"
                    row[loc] = f"{label} ({cell['fte']:.1f})" if cell["fte"] else label
                row["Data Quality"] = matrix["customer_warnings"].get(cust) or "OK"
                matrix_rows.append(row)
            pd.DataFrame(matrix_rows).to_excel(writer, index=False, sheet_name=f"Matrix_{period}_{category}"[:31])

        # --- Location summary ---
        loc_summary = location_summary(dataset.long_df, filters, period, category)
        pd.DataFrame(loc_summary).drop(columns=["customers"], errors="ignore").to_excel(
            writer, index=False, sheet_name="Location Summary")

        # --- Customer summary ---
        cust_summary = customer_summary(dataset.long_df, filters, period, category)
        cust_df = pd.DataFrame(cust_summary)
        if not cust_df.empty:
            cust_df["base_locations"] = cust_df["base_locations"].apply(lambda x: ", ".join(x))
            cust_df["additional_locations"] = cust_df["additional_locations"].apply(lambda x: ", ".join(x))
        cust_df.to_excel(writer, index=False, sheet_name="Customer Summary")

        # --- Period comparison ---
        if from_period and to_period and from_period != to_period:
            comparison = period_comparison(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(comparison).to_excel(writer, index=False, sheet_name="Period Comparison")

            capacity = capacity_view(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(capacity).to_excel(writer, index=False, sheet_name="Capacity View")

            transfers = identify_transfer_opportunities(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(transfers).to_excel(writer, index=False, sheet_name="Transfer Opportunities")

        # --- Manual mappings ---
        pd.DataFrame(manual_mappings or []).to_excel(writer, index=False, sheet_name="Manual Mappings")

        # --- Data quality ---
        dq = data_quality_report(dataset.long_df, dataset.warnings)
        dq_rows = (
            [{"Type": "Parse Warning", "Detail": w} for w in dq["parse_warnings"]]
            + [{"Type": "Missing Base Location", "Detail": c} for c in dq["missing_base_customers"]]
            + [{"Type": "Multiple Base Locations", "Detail": c} for c in dq["multiple_base_customers"]]
            + [{"Type": "Negative FTE", "Detail": str(r)} for r in dq["negative_fte_records"]]
        )
        pd.DataFrame(dq_rows).to_excel(writer, index=False, sheet_name="Data Quality")

        # --- Raw normalized data (reference) ---
        dataset.raw_df.to_excel(writer, index=False, sheet_name="Raw Data"[:31])

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 1 and ws.max_column >= 1:
                _style_header(ws)
                _autosize(ws)

    buf.seek(0)
    return buf.read()

# ============================================================================
# --- from app/assistant_engine.py ---
# ============================================================================
"""
Rule-based "assistant" for the FTE dataset: a curated list of predefined
questions, plus a small query-builder (pick a metric, run it against
whatever filters/period are currently set). Both are answered purely from
analysis_engine/mapping_engine -- there is no external AI/LLM call, so
every answer is exact and reproducible from the same data that powers the
rest of the app.
"""

from typing import Optional

import pandas as pd



class AssistantError(Exception):
    """Raised when a question/metric can't be answered as asked (e.g. missing periods)."""


def _fmt(n) -> str:
    n = round(float(n), 1)
    return str(int(n)) if n == int(n) else str(n)


def _table(columns: list[str], rows: list[list]) -> dict:
    return {"columns": columns, "rows": rows}


def _require_period_pair(from_period: Optional[str], to_period: Optional[str]) -> None:
    if not from_period or not to_period or from_period == to_period:
        raise AssistantError("Pick two different periods (From / To) above, then ask again.")


# ---------------------------------------------------------------------------
# Predefined questions
# ---------------------------------------------------------------------------

PREDEFINED_QUESTIONS = [
    {"id": "total_fte", "label": "What's the total FTE?", "needs": ["period"],
     "keywords": ["total fte", "total", "overall fte", "how much fte"]},
    {"id": "top_location", "label": "Which location has the most FTE?", "needs": ["period"],
     "keywords": ["top location", "most fte location", "highest location", "biggest location"]},
    {"id": "top_customer", "label": "Which customer has the most FTE?", "needs": ["period"],
     "keywords": ["top customer", "most fte customer", "biggest customer", "highest customer"]},
    {"id": "base_issues", "label": "Are there any base-location issues?", "needs": [],
     "keywords": ["base issue", "base location issue", "missing base", "multiple base"]},
    {"id": "negative_fte", "label": "Are there any negative FTE records?", "needs": [],
     "keywords": ["negative fte", "negative", "data quality"]},
    {"id": "releasing_capacity", "label": "Which locations are releasing capacity?", "needs": ["from_period", "to_period"],
     "keywords": ["releasing capacity", "released capacity", "freeing up capacity", "capacity release"]},
    {"id": "growing_locations", "label": "Which locations are growing?", "needs": ["from_period", "to_period"],
     "keywords": ["growing location", "growth", "growing"]},
    {"id": "transfer_opportunities", "label": "What transfer opportunities exist?", "needs": ["from_period", "to_period"],
     "keywords": ["transfer opportunit", "transfer", "matchable"]},
    {"id": "category_split", "label": "What's the Internal / SWC / External / Others split?", "needs": ["period"],
     "keywords": ["category split", "internal swc external", "breakdown", "split"]},
    {"id": "external_dependency", "label": "Which location relies most on External FTE?", "needs": ["period"],
     "keywords": ["external dependency", "external fte", "relies on external", "external reliance"]},
]


def _answer_total_fte(long_df, filters, period, **_):
    filtered = apply_filters(long_df, filters)
    df = fte_for_period_category(filtered, period, "Total")
    return {"summary": f"Total FTE for **{period}**: **{_fmt(df['FTE'].sum())}**.", "table": None}


def _answer_top_location(long_df, filters, period, **_):
    rows = location_summary(long_df, filters, period, "Total")
    if not rows:
        return {"summary": f"No location data for {period} with the current filters.", "table": None}
    top = rows[0]
    return {
        "summary": f"**{top['location']}** has the most FTE in {period}: {_fmt(top['total_fte'])}, "
                   f"supporting {top['customer_count']} customer(s).",
        "table": _table(
            ["Location", "Total FTE", "Customers Supported"],
            [[r["location"], _fmt(r["total_fte"]), r["customer_count"]] for r in rows[:5]],
        ),
    }


def _answer_top_customer(long_df, filters, period, **_):
    rows = customer_summary(long_df, filters, period, "Total")
    if not rows:
        return {"summary": f"No customer data for {period} with the current filters.", "table": None}
    top = rows[0]
    return {
        "summary": f"**{top['customer']}** has the most FTE in {period}: {_fmt(top['total_fte'])}.",
        "table": _table(
            ["Customer", "Total FTE", "Base Location(s)"],
            [[r["customer"], _fmt(r["total_fte"]), ", ".join(r["base_locations"]) or "-"] for r in rows[:5]],
        ),
    }


def _answer_base_issues(long_df, filters, **_):
    filtered = apply_filters(long_df, filters)
    warnings = customer_base_warnings(base_location_status(filtered))
    issues = {c: w for c, w in warnings.items() if w}
    if not issues:
        return {"summary": "No base-location issues found — every customer has exactly one base location.", "table": None}
    return {
        "summary": f"{len(issues)} customer(s) have base-location issues.",
        "table": _table(["Customer", "Issue"], [[c, w] for c, w in sorted(issues.items())]),
    }


def _answer_negative_fte(long_df, filters, **_):
    filtered = apply_filters(long_df, filters)
    negative = filtered[filtered["FTE"] < 0][["Customer_Account", "Location", "Period", "Category", "FTE"]].drop_duplicates()
    if negative.empty:
        return {"summary": "No negative FTE records found.", "table": None}
    return {
        "summary": f"{len(negative)} negative FTE record(s) found — worth a data-quality check.",
        "table": _table(["Customer", "Location", "Period", "Category", "FTE"], negative.values.tolist()),
    }


def _answer_releasing_capacity(long_df, filters, from_period, to_period, **_):
    _require_period_pair(from_period, to_period)
    releasing = [r for r in capacity_view(long_df, filters, from_period, to_period, "Total")
                 if r["status"] == "Potential Released Capacity"]
    if not releasing:
        return {"summary": f"No locations are releasing capacity between {from_period} and {to_period}.", "table": None}
    return {
        "summary": f"{len(releasing)} location(s) are releasing capacity between {from_period} and {to_period}.",
        "table": _table(
            ["Location", "Current FTE", "Future FTE", "Change"],
            [[r["location"], _fmt(r["current_fte"]), _fmt(r["future_fte"]), _fmt(r["change"])] for r in releasing],
        ),
    }


def _answer_growing_locations(long_df, filters, from_period, to_period, **_):
    _require_period_pair(from_period, to_period)
    growing = [r for r in capacity_view(long_df, filters, from_period, to_period, "Total") if r["status"] == "Growth"]
    if not growing:
        return {"summary": f"No locations are growing between {from_period} and {to_period}.", "table": None}
    return {
        "summary": f"{len(growing)} location(s) are growing between {from_period} and {to_period}.",
        "table": _table(
            ["Location", "Current FTE", "Future FTE", "Change"],
            [[r["location"], _fmt(r["current_fte"]), _fmt(r["future_fte"]), _fmt(r["change"])] for r in growing],
        ),
    }


def _answer_transfer_opportunities(long_df, filters, from_period, to_period, **_):
    _require_period_pair(from_period, to_period)
    rows = identify_transfer_opportunities(long_df, filters, from_period, to_period, "Total")
    if not rows:
        return {"summary": f"No transfer opportunities found between {from_period} and {to_period}.", "table": None}
    return {
        "summary": f"{len(rows)} potential transfer opportunit{'y' if len(rows) == 1 else 'ies'} "
                   f"between {from_period} and {to_period}.",
        "table": _table(
            ["From Location", "To Location", "Matchable FTE"],
            [[r["from_location"], r["to_location"], _fmt(r["matchable_fte"])] for r in rows[:10]],
        ),
    }


def _answer_category_split(long_df, filters, period, **_):
    result = category_breakdown(long_df, filters, period)
    if not result["categories"]:
        return {"summary": "This workbook has no Internal/SWC/External/Others breakdown — only period totals.", "table": None}
    total = sum(d["fte"] for d in result["by_category"])
    if total <= 0:
        return {"summary": f"No FTE recorded for {period} with the current filters.", "table": None}
    parts = ", ".join(f"{d['category']} {round(d['fte'] / total * 100)}%" for d in result["by_category"])
    return {
        "summary": f"For {period}, the split is: {parts} (total {_fmt(total)} FTE).",
        "table": _table(
            ["Category", "FTE", "% of Total"],
            [[d["category"], _fmt(d["fte"]), f"{round(d['fte'] / total * 100)}%"] for d in result["by_category"]],
        ),
    }


def _answer_external_dependency(long_df, filters, period, **_):
    result = category_breakdown(long_df, filters, period)
    if "External" not in result["categories"] or not result["by_location"]:
        return {"summary": "This workbook has no External-category data to compare.", "table": None}
    ranked = sorted(
        (r for r in result["by_location"] if r["total"] > 0),
        key=lambda r: r.get("External", 0) / r["total"],
        reverse=True,
    )
    if not ranked:
        return {"summary": f"No FTE recorded for {period} with the current filters.", "table": None}
    top = ranked[0]
    pct = round(top.get("External", 0) / top["total"] * 100)
    return {
        "summary": f"**{top['location']}** relies most on External FTE in {period}: "
                   f"{_fmt(top.get('External', 0))} of {_fmt(top['total'])} total ({pct}%).",
        "table": _table(
            ["Location", "External FTE", "Total FTE", "% External"],
            [[r["location"], _fmt(r.get("External", 0)), _fmt(r["total"]), f"{round(r.get('External', 0) / r['total'] * 100)}%"]
             for r in ranked[:5]],
        ),
    }


_PREDEFINED_HANDLERS = {
    "total_fte": _answer_total_fte,
    "top_location": _answer_top_location,
    "top_customer": _answer_top_customer,
    "base_issues": _answer_base_issues,
    "negative_fte": _answer_negative_fte,
    "releasing_capacity": _answer_releasing_capacity,
    "growing_locations": _answer_growing_locations,
    "transfer_opportunities": _answer_transfer_opportunities,
    "category_split": _answer_category_split,
    "external_dependency": _answer_external_dependency,
}


def answer_predefined(
    question_id: str, long_df: pd.DataFrame, filters: dict,
    period: Optional[str], from_period: Optional[str], to_period: Optional[str],
) -> dict:
    handler = _PREDEFINED_HANDLERS.get(question_id)
    if handler is None:
        raise AssistantError(f"Unknown question: {question_id}")
    return handler(long_df, filters, period=period, from_period=from_period, to_period=to_period)


# ---------------------------------------------------------------------------
# Free-text box: no NLP/LLM involved. Resolves locally, in order, to
# (1) an exact customer or location name, (2) a keyword match against the
# predefined questions above, or (3) a "try a quick question" fallback.
# ---------------------------------------------------------------------------

def _lookup_customer(long_df, filters, period, name):
    rows = customer_summary(long_df, filters, period, "Total")
    match = next((r for r in rows if r["customer"].lower() == name.lower()), None)
    if not match:
        return None
    return {
        "summary": f"**{match['customer']}**: {_fmt(match['total_fte'])} FTE in {period}. "
                   f"Base: {', '.join(match['base_locations']) or '-'}. "
                   f"Additional: {', '.join(match['additional_locations']) or '-'}. "
                   f"Status: {match['warning'] or 'OK'}.",
        "table": None,
    }


def _lookup_location(long_df, filters, period, name):
    rows = location_summary(long_df, filters, period, "Total")
    match = next((r for r in rows if r["location"].lower() == name.lower()), None)
    if not match:
        return None
    return {
        "summary": f"**{match['location']}**: {_fmt(match['total_fte'])} FTE in {period}, "
                   f"supporting {match['customer_count']} customer(s), base location for {match['base_for_count']} of them.",
        "table": None,
    }


def answer_freeform(
    text: str, long_df: pd.DataFrame, filters: dict,
    period: Optional[str], from_period: Optional[str], to_period: Optional[str],
) -> dict:
    q = (text or "").strip().lower()
    if not q:
        raise AssistantError("Type a question, or tap one of the quick options below.")

    filtered = apply_filters(long_df, filters)
    customers = {str(c).lower(): c for c in filtered["Customer_Account"].dropna().unique()}
    locations = {str(l).lower(): l for l in filtered["Location"].dropna().unique()}

    if q in customers and period:
        result = _lookup_customer(long_df, filters, period, customers[q])
        if result:
            return result
    if q in locations and period:
        result = _lookup_location(long_df, filters, period, locations[q])
        if result:
            return result

    best_id, best_score = None, 0
    for item in PREDEFINED_QUESTIONS:
        for kw in item.get("keywords", []):
            if kw in q or q in kw:
                score = len(kw)
                if score > best_score:
                    best_id, best_score = item["id"], score
    if best_id:
        return answer_predefined(best_id, long_df, filters, period, from_period, to_period)

    raise AssistantError(
        "I don't have a canned answer for that yet — try one of the quick questions below, "
        "or type a customer/location name exactly as it appears in your data."
    )


# ---------------------------------------------------------------------------
# Query builder: pick a metric, run it against the current filters/period.
# Same underlying data as the predefined questions, but returns the full
# listing rather than just the top result.
# ---------------------------------------------------------------------------

METRICS = [
    {"id": "total_fte", "label": "Total FTE", "needs": ["period"]},
    {"id": "location_breakdown", "label": "FTE by Location", "needs": ["period"]},
    {"id": "customer_breakdown", "label": "FTE by Customer", "needs": ["period"]},
    {"id": "category_breakdown", "label": "FTE by Category (Internal/SWC/External/Others)", "needs": ["period"]},
    {"id": "base_status", "label": "Base-location status", "needs": []},
    {"id": "period_comparison", "label": "Period-over-period change", "needs": ["from_period", "to_period"]},
    {"id": "transfer_opportunities", "label": "Transfer opportunities", "needs": ["from_period", "to_period"]},
]


def _metric_location_breakdown(long_df, filters, period, **_):
    rows = location_summary(long_df, filters, period, "Total")
    if not rows:
        return {"summary": f"No location data for {period} with the current filters.", "table": None}
    total = sum(r["total_fte"] for r in rows)
    return {
        "summary": f"{len(rows)} location(s), {_fmt(total)} FTE total for {period}.",
        "table": _table(
            ["Location", "Total FTE", "Customers Supported", "Base For"],
            [[r["location"], _fmt(r["total_fte"]), r["customer_count"], r["base_for_count"]] for r in rows],
        ),
    }


def _metric_customer_breakdown(long_df, filters, period, **_):
    rows = customer_summary(long_df, filters, period, "Total")
    if not rows:
        return {"summary": f"No customer data for {period} with the current filters.", "table": None}
    total = sum(r["total_fte"] for r in rows)
    return {
        "summary": f"{len(rows)} customer(s), {_fmt(total)} FTE total for {period}.",
        "table": _table(
            ["Customer", "Total FTE", "Base Location(s)", "Additional Locations", "Status"],
            [[r["customer"], _fmt(r["total_fte"]), ", ".join(r["base_locations"]) or "-",
              ", ".join(r["additional_locations"]) or "-", r["warning"] or "OK"] for r in rows],
        ),
    }


def _metric_base_status(long_df, filters, **_):
    filtered = apply_filters(long_df, filters)
    warnings = customer_base_warnings(base_location_status(filtered))
    issues = sum(1 for w in warnings.values() if w)
    summary = (f"{len(warnings)} customer(s) checked — all OK." if not issues else
               f"{len(warnings)} customer(s) checked; {issues} with a base-location issue.")
    return {
        "summary": summary,
        "table": _table(["Customer", "Status"], [[c, w or "OK"] for c, w in sorted(warnings.items())]),
    }


def _metric_period_comparison(long_df, filters, from_period, to_period, **_):
    _require_period_pair(from_period, to_period)
    rows = period_comparison(long_df, filters, from_period, to_period, "Total")
    if not rows:
        return {"summary": f"No data to compare between {from_period} and {to_period}.", "table": None}
    growth = sum(1 for r in rows if r["status"] == "Growth")
    reduction = sum(1 for r in rows if r["status"] == "Reduction")
    return {
        "summary": f"Between {from_period} and {to_period}: {growth} growing, {reduction} reducing, "
                   f"{len(rows)} customer/location combination(s) total.",
        "table": _table(
            ["Customer", "Location", "From FTE", "To FTE", "Change", "Status"],
            [[r["customer"], r["location"], _fmt(r["from_fte"]), _fmt(r["to_fte"]), _fmt(r["change"]), r["status"]]
             for r in rows[:15]],
        ),
    }


_METRIC_HANDLERS = {
    "total_fte": _answer_total_fte,
    "location_breakdown": _metric_location_breakdown,
    "customer_breakdown": _metric_customer_breakdown,
    "category_breakdown": _answer_category_split,
    "base_status": _metric_base_status,
    "period_comparison": _metric_period_comparison,
    "transfer_opportunities": _answer_transfer_opportunities,
}


def answer_query(
    metric_id: str, long_df: pd.DataFrame, filters: dict,
    period: Optional[str], from_period: Optional[str], to_period: Optional[str],
) -> dict:
    handler = _METRIC_HANDLERS.get(metric_id)
    if handler is None:
        raise AssistantError(f"Unknown metric: {metric_id}")
    return handler(long_df, filters, period=period, from_period=from_period, to_period=to_period)

# ============================================================================
# --- from app/main.py ---
# ============================================================================
import logging
import uuid
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse, Response


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("fte_planner")

import tempfile
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = Path(tempfile.gettempdir()) / "optiview_uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI(title="OptiView — FTE Location & Workforce Planning Tool")
templates = SimpleTemplates(TEMPLATES)

# manual mappings: dataset_id -> list[dict]
_manual_mappings: dict[str, list[dict]] = {}

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
MAX_UPLOAD_MB = 25


def get_dataset(request: Request):
    dataset_id = request.cookies.get(SESSION_COOKIE)
    if not dataset_id:
        return None, None
    dataset = store.get(dataset_id)
    return dataset_id, dataset


def require_dataset(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        raise HTTPException(status_code=400, detail="No dataset uploaded yet")
    return dataset_id, dataset


def parse_filters(request: Request) -> dict:
    q = request.query_params
    filters = {}
    for dim in FILTERABLE_DIMS:
        values = q.getlist(dim)
        values = [v for v in values if v and v != "All"]
        if values:
            filters[dim] = values
    return filters


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    dataset_id, dataset = get_dataset(request)
    return templates.TemplateResponse("index.html", {
        "request": request,
        "has_dataset": dataset is not None,
        "filename": dataset.source_filename if dataset else None,
    })


@app.post("/upload")
async def upload(request: Request, file: UploadFile = File(...)):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "message": "Unsupported file type. Please upload a .xlsx or .xls file.",
        }, status_code=400)

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_MB * 1024 * 1024:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "message": f"File is too large. Maximum size is {MAX_UPLOAD_MB} MB.",
        }, status_code=400)

    tmp_path = UPLOAD_DIR / f"{uuid.uuid4().hex}{suffix}"
    tmp_path.write_bytes(contents)

    try:
        dataset = process_workbook(tmp_path, file.filename)
    except ExcelProcessingError as exc:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "message": str(exc),
            "detail": (
                "Possible reasons: required column is missing, Excel format is invalid, "
                "FTE columns could not be detected, or the file is corrupted."
            ),
        }, status_code=400)
    except Exception:
        logger.exception("Unexpected error processing workbook")
        return templates.TemplateResponse("error.html", {
            "request": request,
            "message": "Unable to process the uploaded Excel file.",
            "detail": "An unexpected error occurred. Please check the file and try again.",
        }, status_code=500)
    finally:
        tmp_path.unlink(missing_ok=True)

    dataset_id = store.new_id()
    store.put(dataset_id, dataset)
    _manual_mappings[dataset_id] = []

    response = RedirectResponse(url="/dashboard", status_code=303)
    response.set_cookie(SESSION_COOKIE, dataset_id, max_age=60 * 60 * 8)
    return response


@app.post("/reset")
async def reset(request: Request):
    dataset_id, _ = get_dataset(request)
    if dataset_id:
        store.delete(dataset_id)
        _manual_mappings.pop(dataset_id, None)
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


def _common_context(request: Request, dataset):
    return {
        "request": request,
        "filename": dataset.source_filename,
        "periods": dataset.periods,
        "categories": dataset.categories,
        "dims": dataset.dims,
        "warnings": dataset.warnings,
    }


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    return templates.TemplateResponse("dashboard.html", ctx)


@app.get("/mapping", response_class=HTMLResponse)
async def mapping_page(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    return templates.TemplateResponse("mapping.html", ctx)


@app.get("/analysis", response_class=HTMLResponse)
async def analysis_page(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    return templates.TemplateResponse("analysis.html", ctx)


@app.get("/transfer", response_class=HTMLResponse)
async def transfer_page(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    return templates.TemplateResponse("transfer.html", ctx)


@app.get("/data-quality", response_class=HTMLResponse)
async def data_quality_page(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    report = data_quality_report(dataset.long_df, dataset.warnings)
    ctx["report"] = report
    return templates.TemplateResponse("data_quality.html", ctx)


@app.get("/management", response_class=HTMLResponse)
async def management_page(request: Request):
    dataset_id, dataset = get_dataset(request)
    if dataset is None:
        return RedirectResponse(url="/")
    ctx = _common_context(request, dataset)
    return templates.TemplateResponse("management.html", ctx)




# ---------------------------------------------------------------------------
# JSON APIs (consumed by static/js/app.js for live filtering)
# ---------------------------------------------------------------------------

@app.get("/api/meta")
async def api_meta(request: Request):
    dataset_id, dataset = require_dataset(request)
    return {
        "periods": dataset.periods,
        "categories": dataset.categories,
        "dims": dataset.dims,
        "warnings": dataset.warnings,
    }


@app.get("/api/matrix")
async def api_matrix(request: Request, period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return build_json_safe(build_matrix(dataset.long_df, filters, period, category))


@app.get("/api/matrix-breakdown")
async def api_matrix_breakdown(request: Request, period: str):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return build_json_safe(build_matrix_breakdown(dataset.long_df, filters, period))


@app.get("/api/category-breakdown")
async def api_category_breakdown(request: Request, period: str):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return build_json_safe(category_breakdown(dataset.long_df, filters, period))


@app.get("/api/summary")
async def api_summary(request: Request, period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return summary_cards(dataset.long_df, filters, period, category)


@app.get("/api/location-summary")
async def api_location_summary(request: Request, period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return location_summary(dataset.long_df, filters, period, category)


@app.get("/api/customer-summary")
async def api_customer_summary(request: Request, period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return customer_summary(dataset.long_df, filters, period, category)


@app.get("/api/comparison")
async def api_comparison(request: Request, from_period: str, to_period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return period_comparison(dataset.long_df, filters, from_period, to_period, category)


@app.get("/api/capacity")
async def api_capacity(request: Request, from_period: str, to_period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return capacity_view(dataset.long_df, filters, from_period, to_period, category)


@app.get("/api/transfer-opportunities")
async def api_transfer(request: Request, from_period: str, to_period: str, category: str = "Total"):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    return identify_transfer_opportunities(dataset.long_df, filters, from_period, to_period, category)


@app.get("/api/assistant/meta")
async def api_assistant_meta(request: Request):
    require_dataset(request)
    return {
        "questions": PREDEFINED_QUESTIONS,
        "metrics": METRICS,
    }


@app.get("/api/assistant/ask")
async def api_assistant_ask(
    request: Request, question_id: str,
    period: Optional[str] = None, from_period: Optional[str] = None, to_period: Optional[str] = None,
):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    try:
        result = answer_predefined(question_id, dataset.long_df, filters, period, from_period, to_period)
    except AssistantError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    return build_json_safe(result)


@app.get("/api/assistant/query")
async def api_assistant_query(
    request: Request, metric_id: str,
    period: Optional[str] = None, from_period: Optional[str] = None, to_period: Optional[str] = None,
):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    try:
        result = answer_query(metric_id, dataset.long_df, filters, period, from_period, to_period)
    except AssistantError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    return build_json_safe(result)


@app.get("/api/assistant/freeform")
async def api_assistant_freeform(
    request: Request, q: str,
    period: Optional[str] = None, from_period: Optional[str] = None, to_period: Optional[str] = None,
):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    try:
        result = answer_freeform(q, dataset.long_df, filters, period, from_period, to_period)
    except AssistantError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    return build_json_safe(result)


@app.get("/api/manual-mapping")
async def list_manual_mapping(request: Request):
    dataset_id, dataset = require_dataset(request)
    return _manual_mappings.get(dataset_id, [])


@app.post("/api/manual-mapping")
async def add_manual_mapping(
    request: Request,
    from_customer: str = Form(...),
    from_location: str = Form(...),
    to_customer: str = Form(...),
    to_location: str = Form(...),
    fte_amount: float = Form(...),
    notes: str = Form(""),
):
    dataset_id, dataset = require_dataset(request)
    entry = {
        "id": uuid.uuid4().hex[:8],
        "from_customer": from_customer,
        "from_location": from_location,
        "to_customer": to_customer,
        "to_location": to_location,
        "fte_amount": fte_amount,
        "notes": notes,
    }
    _manual_mappings.setdefault(dataset_id, []).append(entry)
    return entry


@app.delete("/api/manual-mapping/{mapping_id}")
async def delete_manual_mapping(request: Request, mapping_id: str):
    dataset_id, dataset = require_dataset(request)
    mappings = _manual_mappings.get(dataset_id, [])
    _manual_mappings[dataset_id] = [m for m in mappings if m["id"] != mapping_id]
    return {"ok": True}


@app.get("/export")
async def export(
    request: Request,
    period: str,
    category: str = "Total",
    from_period: Optional[str] = None,
    to_period: Optional[str] = None,
    view: Optional[str] = None,
):
    dataset_id, dataset = require_dataset(request)
    filters = parse_filters(request)
    manual_mappings = _manual_mappings.get(dataset_id, [])
    content = build_export_workbook(
        dataset, filters, period, category, from_period, to_period, manual_mappings, view=view
    )
    filename = "OptiView_FTE_Analysis.xlsx"
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def build_json_safe(obj):
    """Recursively convert numpy/pandas scalar types to plain Python for JSON."""
    import numpy as np
    if isinstance(obj, dict):
        return {k: build_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [build_json_safe(v) for v in obj]
    if isinstance(obj, (np.generic,)):
        return obj.item()
    return obj


@app.exception_handler(404)
async def not_found(request: Request, exc):
    return templates.TemplateResponse("error.html", {
        "request": request, "message": "Page not found.", "detail": None,
    }, status_code=404)


@app.get("/static/css/style.css")
async def _static_style_css():
    return Response(content=STYLE_CSS, media_type="text/css")


@app.get("/static/js/app.js")
async def _static_app_js():
    return Response(content=APP_JS, media_type="application/javascript")


if __name__ == "__main__":
    import threading
    import webbrowser
    import uvicorn

    URL = "http://127.0.0.1:8000"

    def _open_browser():
        try:
            webbrowser.open(URL)
        except Exception:
            pass  # headless machine or no default browser -- user can open the URL manually

    threading.Timer(1.25, _open_browser).start()
    print(f"OptiView starting -- opening {URL} in your browser...")
    print("(If it doesn't open automatically, copy that URL into your browser.)")
    print("Press CTRL+C to stop.")
    uvicorn.run(app, host="127.0.0.1", port=8000)
