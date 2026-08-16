import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd
import pytest

from app.excel_processor import detect_columns, process_workbook, ExcelProcessingError
from app.utils import parse_fte_column, normalize_base_flag


def test_parse_fte_column_with_category():
    period, category, sort_key = parse_fte_column("FTE_Dec_2025_Internal")
    assert period == "Dec-2025"
    assert category == "Internal"


def test_parse_fte_column_total_only():
    period, category, sort_key = parse_fte_column("FTE_Mar_2026")
    assert period == "Mar-2026"
    assert category == "Total"


def test_parse_fte_column_rejects_unrelated():
    assert parse_fte_column("Customer Account") is None


def test_normalize_base_flag_variants():
    assert normalize_base_flag("Yes") is True
    assert normalize_base_flag("Y") is True
    assert normalize_base_flag("1") is True
    assert normalize_base_flag("TRUE") is True
    assert normalize_base_flag("No") is False
    assert normalize_base_flag("N") is False
    assert normalize_base_flag(None) is None


def test_detect_columns_flexible_names():
    df = pd.DataFrame({
        "S.No": [1],
        "Customer Account": ["Ford"],
        "Location": ["Cob"],
        "Base location": ["Yes"],
        "FTE_Dec_2025_Internal": [5],
        "FTE_Dec_2025": [5],
    })
    detected = detect_columns(df)
    assert detected.customer_account == "Customer Account"
    assert detected.location == "Location"
    assert not detected.missing_required
    assert "Dec-2025" in detected.fte_columns


def test_process_workbook_missing_required_column(tmp_path):
    df = pd.DataFrame({"Foo": [1, 2], "Bar": [3, 4]})
    path = tmp_path / "bad.xlsx"
    df.to_excel(path, index=False)
    with pytest.raises(ExcelProcessingError):
        process_workbook(path, "bad.xlsx")


def test_process_workbook_two_row_header_format(tmp_path):
    """The standard 'FTE_Location_Planner' template: row 1 has the period
    label merged across 5 columns, row 2 has Internal/SWC/External/Others
    and the repeated period total."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FTE_Data"

    headers_row1 = [
        "S.No", "VM Product", "Customer Account", "Component", "Country",
        "Location", "Base location", "FTE_Dec_2025", None, None, None, None,
    ]
    headers_row2 = [
        None, None, None, None, None, None, None,
        "Internal", "SWC", "External", "Others", "FTE_Dec_2025",
    ]
    ws.append(headers_row1)
    ws.append(headers_row2)
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:D2")
    ws.merge_cells("E1:E2")
    ws.merge_cells("F1:F2")
    ws.merge_cells("G1:G2")
    ws.merge_cells("H1:L1")
    ws.append([1, "BS/OSS/ST", "Ford", "NET/DCOM", "India", "Ban", "No", 10, 5, 3, 0, 18])
    ws.append([2, "BS/OSS/ST", "Ford", "NET/DCOM", "India", "Cob", "Yes", 12, 7, 5, 0, 24])

    path = tmp_path / "two_row_header.xlsx"
    wb.save(path)

    dataset = process_workbook(path, "two_row_header.xlsx")
    assert dataset.periods == ["Dec-2025"]
    assert set(dataset.categories) == {"Internal", "SWC", "External", "Others", "Total"}
    assert set(dataset.dims["Customer_Account"]) == {"Ford"}

    totals = dataset.long_df[
        (dataset.long_df["Category"] == "Total") & (dataset.long_df["Location"] == "Ban")
    ]
    assert totals["FTE"].iloc[0] == 18


def test_process_workbook_end_to_end(tmp_path):
    df = pd.DataFrame({
        "S.No": [1, 2, 3],
        "Customer Account": ["Ford", "Ford", "Toyota"],
        "Location": ["Cob", "Ban", "Cob"],
        "Base location": ["Yes", "No", "No"],
        "FTE_Dec_2025_Internal": [10, 5, 3],
        "FTE_Dec_2025_SWC": [1, 1, 1],
        "FTE_Dec_2025": [11, 6, 4],
    })
    path = tmp_path / "good.xlsx"
    df.to_excel(path, index=False)
    dataset = process_workbook(path, "good.xlsx")
    assert dataset.periods == ["Dec-2025"]
    assert "Total" in dataset.categories
    assert "Internal" in dataset.categories
    assert set(dataset.dims["Customer_Account"]) == {"Ford", "Toyota"}
