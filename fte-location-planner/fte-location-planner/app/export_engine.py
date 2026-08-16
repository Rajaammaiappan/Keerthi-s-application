"""
Builds a multi-sheet analyzed Excel workbook the user can download,
covering the matrix, summaries, comparison, capacity view, transfer
opportunities, manual mappings, and data-quality findings.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.mapping_engine import build_matrix
from app import analysis_engine as ae

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BASE_FILL = PatternFill("solid", fgColor="FFE699")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_export_workbook(dataset, filters: dict, period: str, category: str,
                           from_period: str, to_period: str,
                           manual_mappings: list[dict]) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # --- Matrix sheet ---
        matrix = build_matrix(dataset.long_df, filters, period, category)
        matrix_rows = []
        for cust in matrix["customers"]:
            row = {"Customer Account": cust}
            for loc in matrix["locations"]:
                cell = matrix["cells"][cust][loc]
                if cell["is_base"]:
                    label = "BASE"
                elif cell["has_fte"]:
                    label = "SUPPORT"
                else:
                    label = "-"
                row[loc] = f"{label} ({cell['fte']:.1f})" if cell["fte"] else label
            row["Data Quality"] = matrix["customer_warnings"].get(cust) or "OK"
            matrix_rows.append(row)
        pd.DataFrame(matrix_rows).to_excel(writer, index=False, sheet_name=f"Matrix_{period}_{category}"[:31])

        # --- Location summary ---
        loc_summary = ae.location_summary(dataset.long_df, filters, period, category)
        pd.DataFrame(loc_summary).drop(columns=["customers"], errors="ignore").to_excel(
            writer, index=False, sheet_name="Location Summary")

        # --- Customer summary ---
        cust_summary = ae.customer_summary(dataset.long_df, filters, period, category)
        cust_df = pd.DataFrame(cust_summary)
        if not cust_df.empty:
            cust_df["base_locations"] = cust_df["base_locations"].apply(lambda x: ", ".join(x))
            cust_df["additional_locations"] = cust_df["additional_locations"].apply(lambda x: ", ".join(x))
        cust_df.to_excel(writer, index=False, sheet_name="Customer Summary")

        # --- Period comparison ---
        if from_period and to_period and from_period != to_period:
            comparison = ae.period_comparison(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(comparison).to_excel(writer, index=False, sheet_name="Period Comparison")

            capacity = ae.capacity_view(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(capacity).to_excel(writer, index=False, sheet_name="Capacity View")

            transfers = ae.identify_transfer_opportunities(dataset.long_df, filters, from_period, to_period, category)
            pd.DataFrame(transfers).to_excel(writer, index=False, sheet_name="Transfer Opportunities")

        # --- Manual mappings ---
        pd.DataFrame(manual_mappings or []).to_excel(writer, index=False, sheet_name="Manual Mappings")

        # --- Data quality ---
        dq = ae.data_quality_report(dataset.long_df, dataset.warnings)
        dq_rows = (
            [{"Type": "Parse Warning", "Detail": w} for w in dq["parse_warnings"]]
            + [{"Type": "Missing Base Location", "Detail": c} for c in dq["missing_base_customers"]]
            + [{"Type": "Multiple Base Locations", "Detail": c} for c in dq["multiple_base_customers"]]
            + [{"Type": "Negative FTE", "Detail": str(r)} for r in dq["negative_fte_records"]]
        )
        pd.DataFrame(dq_rows).to_excel(writer, index=False, sheet_name="Data Quality")

        # --- Raw normalized data (reference) ---
        dataset.raw_df.to_excel(writer, index=False, sheet_name="Raw Data"[:31])

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 1 and ws.max_column >= 1:
                _style_header(ws)
                _autosize(ws)

    buf.seek(0)
    return buf.read()
